# backend/app/services/ingest.py
# ------------------------------------------------------------
# CellTrail Ingest Service
# - 支援 CSV / TXT / XLSX / PDF
# - PDF 直接在後端解析（手機用戶無需先轉檔）
# - 欄位對照：時間、地址、cell_id、方位角等
# - 缺值/#N/A 正規化、時間標準化 (UTC+8)
# - 先以 cell_id 查站點字典，失敗再用地址地理編碼
# - DB 寫入時自動帶入 geom；避免 server-side prepared 造成錯誤
# ------------------------------------------------------------

from __future__ import annotations
import csv, io, os, re
from datetime import datetime, timezone, timedelta
from typing import Dict, Any, List, Optional, Iterable, Tuple

from fastapi import HTTPException
from app.db.session import get_conn
from app.services import geocode

# ====== 共用常數與工具 ======
NA_TOKENS = {"#N/A", "", "NA", "N/A", None}
TPE_TZ = timezone(timedelta(hours=8))  # Asia/Taipei

# OLE2 / CDFV2 compound document magic：密碼保護的 OOXML（加密 xlsx）會把
# 真正的 zip 包在 OLE2 容器的 EncryptedPackage 串流裡，檔頭即為此 8 bytes。
# （一般未加密的 xlsx 是 zip，檔頭為 PK\x03\x04，不會命中。）
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class EncryptedFileError(Exception):
    """上傳的檔案為密碼保護（加密）的 Office 檔，系統無法解析。

    刻意不做解密：要求使用者先在 Excel 移除密碼、另存為一般 .xlsx 再上傳
    （證據完整性考量 —— 系統不持有/處理密碼，避免成為密碼處理的一環）。
    """
    pass


def _reject_if_encrypted(file_bytes: bytes) -> None:
    """
    上傳前置檢查：若檔案是加密 / 密碼保護的 Office 檔（OLE2 容器），直接
    拋 EncryptedFileError，由端點轉成清楚的錯誤提醒，而非讓 openpyxl 在
    下游拋出難懂的「File is not a zip file」。
    """
    if file_bytes[:8] == _OLE2_MAGIC:
        raise EncryptedFileError(
            "此檔案有密碼保護（加密），系統無法讀取。"
            "請在 Excel 中移除密碼後，另存為一般 .xlsx 再重新上傳。"
        )


class ParseDiagnosisError(Exception):
    """
    解析失敗時拋出，附帶診斷資訊供前端「智慧錯誤診斷」UI 使用。

    diagnosis dict 結構：
      {
        "found_time_col":     bool,
        "found_time_col_name": str | None,
        "found_cell_id_col":  bool,
        "found_cell_id_col_name": str | None,
        "found_addr_col":     bool,
        "found_addr_col_name": str | None,
        "available_columns":  list[str],
      }
    """
    def __init__(self, message: str, diagnosis: Dict[str, Any]):
        super().__init__(message)
        self.diagnosis = diagnosis

def _is_na(v):
    return v is None or (isinstance(v, str) and v.strip() in NA_TOKENS)

def _parse_ts(s: Any) -> Optional[datetime]:
    """
    支援多種時間表示，回傳含台北時區的 datetime：
      - 2025/8/30 13:31         （PDF 漫遊紀錄常見：單位數月日）
      - 2025/08/30 13:31:22     （CSV 標準格式）
      - 2024-09-01 20:06:44     （Excel 網路歷程：dash 連字符）
      - 2024-09-01\xa020:06:44  （Excel 拷貝出來常帶不間斷空格 NBSP）
      - 2023-01-12T00:48:02.000 （W2.4：中華上網方言 ISO 8601 + 毫秒）
      - 2023-01-12T00:48:02     （W2.4：ISO 8601 無毫秒）
      - 中文「年月日時分秒」夾雜（最後備援）

    W2.4 設計筆記（為什麼這樣加 ISO 8601）：
      中華上網方言把時間放在「起台」欄、格式是 ISO 8601 帶 T 分隔 + .000 毫秒。
      這個格式跟現有「網路歷程.xltx」的 dash+space 格式語意完全等價、只是分隔符
      不同，所以用「加新 fmt 進 list」的方式擴充最安全 —— 對其他方言的 row 不會
      造成誤判（Python strptime 嚴格匹配，T 分隔不會 match dash+space 格式）。

      不用 datetime.fromisoformat 是因為：
        (1) Python 3.10 之前不接受 'Z' 後綴，行為跨版本不一致；
        (2) 它會吃 '+08:00' 時區資訊，可能跟我們手動加的 TPE_TZ 衝突；
        (3) 我們刻意只接受「無時區後綴」的 naïve 字串，假設都是台北時間。
    """
    if _is_na(s):
        return None
    if isinstance(s, datetime):
        return s if s.tzinfo else s.replace(tzinfo=TPE_TZ)
    # NBSP（不間斷空格 \xa0）→ 一般空格；前後 strip
    s = str(s).strip().replace("\xa0", " ")
    # 把多重空白壓成單一空白（Excel 偶爾會有兩個空格）
    s = re.sub(r"[ \t]+", " ", s)
    for fmt in (
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%-m/%-d %H:%M",
        "%Y-%m-%d %H:%M:%S",  # 「網路歷程.xltx」格式
        "%Y-%m-%d %H:%M",
        # ── W2.4：ISO 8601（T 分隔）─────────────────────────────
        # 順序：先試含毫秒（更嚴格的格式優先），strptime 嚴格匹配所以
        # "2023-01-12T00:48:02.000" 不會誤 match "%Y-%m-%dT%H:%M:%S"
        "%Y-%m-%dT%H:%M:%S.%f",  # W2.4：中華上網方言 12869 列實測 100% 命中
        "%Y-%m-%dT%H:%M:%S",     # W2.4：ISO 8601 無毫秒備援（其他 carrier 可能用）
        # ── GPS 軌跡格式：M/D/YYYY 12 小時制 + AM/PM（如 "4/18/2026 3:51:09 AM"）──
        # 來源：車機 / GPS 軌跡匯出（RFL-8271軌跡.xlsx）。%m/%d/%I 在 CPython
        # strptime 接受非零填補（"4" 等同 "04"），故單位數月/日/時皆可命中。
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %I:%M %p",
    ):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.replace(tzinfo=TPE_TZ)
        except Exception:
            continue
    # 嘗試去除多餘空白或中文標點
    s2 = re.sub(r"[年月日時分秒]", " ", s)
    s2 = re.sub(r"\s+", " ", s2).strip(" /-:")
    for fmt in ("%Y %m %d %H %M %S", "%Y %m %d %H %M"):
        try:
            dt = datetime.strptime(s2, fmt)
            return dt.replace(tzinfo=TPE_TZ)
        except Exception:
            pass
    return None

def _to_int(s: Any) -> Optional[int]:
    if _is_na(s):
        return None
    s = re.sub(r"[^\d-]", "", str(s).strip())
    if not s:
        return None
    try:
        return int(s)
    except Exception:
        return None

def _to_float(s: Any) -> Optional[float]:
    if _is_na(s):
        return None
    try:
        return float(str(s).strip())
    except Exception:
        return None

def _guess_accuracy(addr: str | None) -> int:
    """簡單估計誤差圈：市區較小、郊區較大"""
    a = (addr or "")
    if ("市" in a) or ("區" in a):
        return 150
    if ("鄉" in a) or ("村" in a):
        return 800
    return 300

# GPS / 經緯度直給座標的誤差圈：車機定位精度通常 < 30m，遠小於基地台
GPS_ACCURACY_M = 30

def _resolve_latlng(r: Dict[str, Any]) -> Optional[Tuple[float, float]]:
    """
    從 normalize 後的 row 取「檔案自帶」的經緯度（GPS 軌跡 / 已含座標格式）。

    回傳 (lat, lng) 或 None（無有效座標）。

    穩健性設計（為什麼不單純信任標頭）：
      實務上電信 / 車機匯出檔常把「經度」「緯度」欄位標反（值對調）。
      例 RFL-8271軌跡.xlsx：標「經度」的欄裝 22.59、標「緯度」的欄裝 120.32，
      但台灣 lat≈22-25、lng≈120-122 —— 緯度物理上必落在 [-90, 90]。
      故以「緯度必在 [-90,90]、經度必在 [-180,180]」做範圍校正：
        - 若 lat 超出 [-90,90] 而 lng 在 [-90,90] → 判定標反，對調。
      仍無法落入合理範圍（兩者皆超界 / 皆 0）→ 視為無效座標回 None，
      讓該列走原本的 cell_id/addr geocode 路徑（forensic「不亂猜」紀律）。
    """
    lat = _to_float(r.get("lat"))
    lng = _to_float(r.get("lng"))
    if lat is None or lng is None:
        return None
    # 範圍自動校正：緯度必在 [-90,90]
    if abs(lat) > 90 and abs(lng) <= 90:
        lat, lng = lng, lat
    if abs(lat) > 90 or abs(lng) > 180:
        return None
    if lat == 0 and lng == 0:
        return None
    return (lat, lng)

# ---- 欄位名標準化 ----
def _canon(s: str) -> str:
    """標準化字串：全形轉半形、移除空白/標點、繁簡（臺→台）、小寫化"""
    if s is None:
        return ""
    s = str(s)
    try:
        import unicodedata
        s = unicodedata.normalize("NFKC", s)
    except Exception:
        pass
    s = s.replace("臺", "台")
    s = re.sub(r"[\s\u3000]+", "", s)  # 各種空白
    s = re.sub(r"[•·．\.\-:：;；,/，、\t]", "", s)  # 常見標點
    return s.lower()

# ====== 讀取 CSV / Excel ======
def _iter_rows_csv(file_bytes: bytes) -> Iterable[Dict[str, Any]]:
    text = file_bytes.decode("utf-8-sig", errors="ignore")
    rdr = csv.DictReader(io.StringIO(text))
    for r in rdr:
        yield {(k or "").strip(): (v.strip() if isinstance(v, str) else v) for k, v in (r or {}).items()}

def _iter_rows_excel(
    file_bytes: bytes,
    user_mapping: Optional[Dict[str, str]] = None,
) -> Iterable[Dict[str, Any]]:
    """
    讀取 .xlsx / .xltx / .xlsm / .xltm 為 row dict。

    手動對應（2026-06-27）：`user_mapping`（{raw_col: system_field}）為使用者在
    「系統不認識此格式」時手動指定的欄位對應。傳入時，header detection 規則 B
    計分會把「使用者已指定的欄位」也算命中 —— 否則完全陌生的欄名（0 個已知
    別名）會在規則 B 被整張 sheet 丟棄，輪不到後續 _apply_user_mapping rename，
    手動對應形同虛設（這正是先前的結構性 bug）。raw key 照舊原樣 yield，rename
    由呼叫端的 _apply_user_mapping 負責，分工不變。

    W2.1（2026-04-29）：多 sheet 支援
    ──────────────────────────────────────────
    舊版只讀第一張表 → 周蔓達 13 月只看到 1 月、電話通聯+歷程 6 個案件
    主體只看到第一個。改為遍歷所有 sheet。

    W2.2（2026-04-29）：表頭埋深偵測 + 命中分數機制
    ────────────────────────────────────────────────
    舊版只能偵測「row 0 大標、row 1 真表頭」這種固定 supertitle pattern，
    對於電話通聯+歷程.xlsx 的真實情境完全失靈：
      - 嫌1 雙向歷程：真表頭在 row 10（前面 5-9 是 PII 個人資料區）
      - 嫌1 網路歷程：真表頭在 row 22（前面 21 列是查詢條件 + 用戶資訊）
      - 嫌2/害 雙向 / 網路歷程：真表頭都在 row 5（前 4 列是查詢條件）

    新版演算法（取代舊 Unnamed 偵測，後者是本演算法 N=2 的特例）：
      1. 用 header=None 讀整張 sheet 為 df_raw
      2. scan 前 N=25 列，逐列計算「該列有幾個 cell 命中 active header_map」
      3. 取「命中數最多」的列當 header（同分時取首次出現）
      4. 若最高命中數 < M=2 → 規則 B 跳過（保留 W2.1 行為的概念）
      5. 若 sheet 總列數 < 5 → 規則 A 跳過（同上）

    為何 N=30：實測最深的真表頭在 row 27（台哥大上網歷程 test2.xlsx，前面是
              查詢條件 + 完整「使用者資料」PII 區塊），預留 buffer。
    為何 M=2：M=1 太鬆（PII sheet 的「地址」單欄會誤命中），
              M=3 太嚴（縮減 schema 會誤殺）。
    為何「首次出現」而非「最後出現」：嫌1 雙向歷程 row 10/11 重複表頭
              文字，挑首次（row 10）→ row 11 變假資料列、ingest 端
              `_parse_ts('始話時間')` 必定失敗 → 自動過濾。雙重保險。

    附帶效益（forensic data minimization）：表頭之上的 PII metadata
    （姓名/身分證/出生）會被自動切掉，不會 yield 出來；純人資 sheet
    （無真表頭可命中）也仍被規則 B 擋下。

    處理電信公司常見的「假表頭」格式：本演算法已自然涵蓋
      - row 0 = 真表頭（純資料表）→ best_idx=0
      - row 0 = 跨欄大標、row 1 = 真表頭（W2.1 supertitle）→ best_idx=1
      - row 0~k = metadata、row k+1 = 真表頭（W2.2 buried）→ best_idx=k+1

    型別處理盲點：
      - pandas.Timestamp / numpy.datetime64 → 用 .to_pydatetime() 轉成
        python datetime（切忌用 .item()，部分版本會回 int [ns]，後續
        _parse_ts 全部失敗）
      - 其他 numpy 數值型別才用 .item() 轉原生
    """
    try:
        import pandas as pd
        import numpy as np
    except Exception as e:
        raise RuntimeError("請先安裝：pip install pandas openpyxl") from e

    # 一次取得 active header_map（W1 架構，DB 為 SoT）；service 內部會
    # 自動 fallback 到 _RAW2CANON。連 service import 都失敗（極端狀況，
    # 例如測試環境完全無 app context）才退到本檔常數
    try:
        from app.services.carrier_profile import get_active_header_map
        active_map = get_active_header_map()
    except Exception:
        active_map = HEADER_MAP

    SCAN_WINDOW = 30         # 表頭最多埋多深（台哥大上網歷程 test2.xlsx 真表頭在
                             # row 27：前面有查詢條件 + 完整「使用者資料」PII 區塊；
                             # 舊上限 25 會掃不到。放寬安全：PII/metadata 列命中數=0，
                             # 真表頭（≥2 canonical 命中）仍穩定勝出，規則 B 不受影響）
    MIN_HEADER_MATCHES = 2   # 真表頭至少要命中幾欄才算數

    # 手動對應：使用者指定的 raw 欄名（canon 後）視同「已知欄位」參與規則 B 計分
    mapped_canon = {
        _canon(str(k))
        for k, v in (user_mapping or {}).items()
        if v and v != "ignore"
    }

    def _is_empty_cell(c) -> bool:
        if c is None:
            return True
        if isinstance(c, str):
            return not c.strip()
        try:
            return bool(pd.isna(c))
        except Exception:
            return False

    # 用 ExcelFile 列出 sheet 名，避免每張都重新解析整個 workbook
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    skipped_sheets: List[Tuple[str, str]] = []  # (sheet_name, reason)

    for sheet_name in xls.sheet_names:
        # 整張 raw 讀（無 header），讓我們可以掃任意列當 header
        df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None)

        # 規則 A：總列數 < 2（連「1 表頭 + 1 資料」都湊不齊）才當非資料表。
        # 為什麼從舊的 < 5 放寬到 < 2（2026-06-22）：
        #   偵查實務上「某對象在調閱期間只有 1～2 筆通聯」是真實且關鍵的資料，
        #   舊門檻會把這種小檔整個跳過 → 沉默丟證據，違反證據完整性原則。
        #   擋「封面頁／統計頁／說明頁」這類非資料 sheet 的真正防線是規則 B
        #   （表頭必須命中 ≥2 個 canonical 別名）＋ line 350（去表頭後須 ≥1 列），
        #   不是列數本身；故安全地把列數門檻降到「資料表的物理下限」= 2 列。
        if len(df_raw) < 2:
            skipped_sheets.append((sheet_name, f"row<2 ({len(df_raw)} rows)"))
            continue

        # W2.2 核心：scan 前 SCAN_WINDOW 列，找命中 header_map 最多的列
        scan_n = min(SCAN_WINDOW, len(df_raw))
        best_idx = -1
        best_match = 0
        for ri in range(scan_n):
            row = df_raw.iloc[ri]
            n = 0
            for c in row:
                if _is_empty_cell(c):
                    continue
                canon = _canon(str(c))
                if active_map.get(canon) or canon in mapped_canon:
                    n += 1
            if n > best_match:
                best_match = n
                best_idx = ri

        # 規則 B：scan 視窗內無夠強的 header → 視為非資料 sheet
        if best_match < MIN_HEADER_MATCHES:
            skipped_sheets.append(
                (sheet_name, f"header matches < {MIN_HEADER_MATCHES} (best={best_match})")
            )
            continue

        # 切片：best_idx 是 header row，其下是資料
        header_row = df_raw.iloc[best_idx]
        header: List[str] = []
        for i, c in enumerate(header_row):
            if _is_empty_cell(c):
                # W2.6：合併表頭偵測 — 若前欄的 canonical mapping 是 cell_id 類，
                # 當前空欄是「基地台/交換機」合併表頭的右半欄（含基地台地址）。
                # 根因：雙向通聯格式 H1:I1 合併儲存格，H=數字 cell_id，I=地址，
                # pandas 讀 I1=None，沒名字的欄無法 normalize 到 cell_addr。
                # 修法：偵測到這個 pattern 就直接補上已知別名「基地台地址」。
                if i > 0 and active_map.get(_canon(header[-1])) in ("cell_id", "cell_id_compound"):
                    header.append("基地台地址")
                else:
                    header.append(f"_unnamed_{i}")
            else:
                header.append(str(c).strip())
        df = df_raw.iloc[best_idx + 1:].copy()
        df.columns = header

        # 規則 A 二次校驗：去 header 後資料列若 < 1，跳過（罕見但保險）
        if len(df) < 1:
            skipped_sheets.append(
                (sheet_name, f"no data row after header at row{best_idx+1}")
            )
            continue

        df = df.replace({np.nan: ""})

        # ── W2.4：dialect detection（per-sheet）──────────────────
        # 在 sheet 進入 yield loop 前一次決定 dialect，避免每 row 重算。
        # 抽樣前 20 row 給 detector，足以判斷主流類別、又不會讀爆大檔。
        sample_rows: List[Dict[str, Any]] = []
        for _, sample_row in df.head(20).iterrows():
            sample_rows.append({str(k).strip(): sample_row[k] for k in df.columns})
        sheet_dialect = _detect_dialect(header, sample_rows)
        # 落 log（forensic 系統應可追溯每 sheet 走的 normalize path）
        if sheet_dialect:
            import logging
            logging.getLogger(__name__).info(
                "ingest: sheet=%r detected dialect=%r", sheet_name, sheet_dialect
            )

        for _, row in df.iterrows():
            d = {str(k).strip(): row[k] for k in df.columns}
            for k, v in list(d.items()):
                # pandas.Timestamp / numpy.datetime64 → python datetime（保留時區邏輯由 _parse_ts 處理）
                if hasattr(v, "to_pydatetime"):
                    try:
                        d[k] = v.to_pydatetime()
                        continue
                    except Exception:
                        pass
                # 其他 numpy 型別 → python 原生（跳過 str/bytes，避免它們意外實作了 .item()）
                try:
                    if hasattr(v, "item") and not isinstance(v, (str, bytes)):
                        d[k] = v.item()
                except Exception:
                    pass
            # W2.4：用 reserved key 注入 dialect tag。下游 _normalize_row 會
            # pop 掉這個 key、決定走 dialect path 或標準 path。命名加雙底線
            # 前後綴是 magic key 的常見 convention，幾乎不可能撞到真欄名。
            if sheet_dialect:
                d["__celltrail_dialect__"] = sheet_dialect
            yield d

    # 跳過的 sheet 留下軌跡（forensic 系統應該可追溯）
    if skipped_sheets:
        import logging
        logging.getLogger(__name__).info(
            "ingest: skipped %d sheet(s): %s",
            len(skipped_sheets),
            ", ".join(f"{n!r}({r})" for n, r in skipped_sheets),
        )

# ====== 欄位對照（來源→系統） ======
# ----------------------------------------------------------------------
# W1 重構（2026-04-27）：
#   原本 _RAW2CANON 是熱路徑用的 source of truth。重構後：
#     - DB 表 carrier_profiles 是新的 SoT；ingest 透過 carrier_profile service 取
#     - 此處 _RAW2CANON 仍保留，但**降級為「DB 不可用時的 fallback 種子」**，
#       同時是單元測試的對照基準（保證 DB 種子與 code fallback 不會飄離）
#
# 加 / 改別名的標準流程（從這個版本起）：
#   1. 直接在 carrier_profiles 表新增 / 修改 mapping_json（Web admin 介面或 SQL）
#   2. 呼叫 carrier_profile.invalidate_cache() 讓 worker 立刻生效
#   3. 不需要改本檔、不需要 deploy
# ----------------------------------------------------------------------
_RAW2CANON = {
    # 時間
    "開始連線時間": "start_ts",
    "結束連線時間": "end_ts",
    "開始時間": "start_ts",
    "結束時間": "end_ts",
    "起始時間": "start_ts",
    "啟始時間": "start_ts",     # 「網路歷程.xltx」用「啟」非「起」
    "終止時間": "end_ts",
    "時間": "start_ts",          # W1 新增：「0801-0903彭奕翔網路歷程.xlsx」
    "始話時間": "start_ts",      # W1 新增：「電話通聯+歷程.xlsx」
    "始話日期時間": "start_ts",  # 雙向通聯：「11501-11505(雙向).xlsx」始話日期+時間合併欄
    "進入基地台時間": "start_ts",  # 台哥大上網歷程「test2.xlsx」：到達該基地台覆蓋的時間（地圖以此為定位時間）
    "離開基地台時間": "end_ts",    # 同上：離開該基地台的時間，補滿 end_ts 不遺失
    "通聯時間": "start_ts",      # W1 新增：常見通聯紀錄欄名
    "手機連到基地台的時間": "start_ts",  # W2.2：「電話通聯+歷程.xlsx」網路歷程 sheet 方言（此 carrier 常空，留作保險）
    "連到internet的時間":   "start_ts",  # W2.2：同上，此 carrier 真正帶值的時間欄
    # 基地台/地址/識別
    "基地台地址": "cell_addr",
    "基地臺地址": "cell_addr",
    "基地台位址": "cell_addr",   # 「網路歷程.xltx」用「位址」非「地址」
    "基地臺位址": "cell_addr",
    "最終基地台位址": "cell_addr",
    "最終基地臺位址": "cell_addr",
    "站台地址": "cell_addr",
    "地址": "cell_addr",
    "起址": "cell_addr",          # W1 新增：「周蔓達上網歷程.xlsx」起話端位址
    "離開基地台地址": "cell_addr",  # 台哥大上網歷程「test2.xlsx」：服務基地台地址
    "基地台編號": "cell_id",
    "基地臺編號": "cell_id",
    "基地台ID": "cell_id",       # Excel 範本有時用 ID 而非「編號」
    "基地臺ID": "cell_id",
    "最終基地台ID": "cell_id",
    "最終基地臺ID": "cell_id",
    "站台編號": "cell_id",
    "站碼": "cell_id",
    "離開基地台編號": "cell_id",  # 台哥大上網歷程「test2.xlsx」：服務基地台編號（純 ID，非複合欄）
    "cell_id": "cell_id",
    "基地台": "cell_id",          # W1 新增：「0801-0903彭奕翔網路歷程.xlsx」
    "基地台/交換機": "cell_id",   # W1 新增：「電話通聯+歷程.xlsx」
    "起台": "cell_id",            # W1 新增：「周蔓達上網歷程.xlsx」起話端
    "基地台代碼": "cell_id",       # W2.2：「電話通聯+歷程.xlsx」網路歷程 sheet 方言
    # ── W2.3：複合欄（cell_id + 空格 + 地址 + 可選代次標籤）──
    # canonical key 故意用 cell_id_compound 而非 cell_id，作為 dispatch tag：
    # _normalize_row 看到此 key 會走「拆解 → 分填 cell_id / cell_addr」路徑。
    # 直接 mapping 到 cell_id 會誤觸其他 carrier 的純 ID 欄，破壞既有行為。
    "迄基地台":   "cell_id_compound",   # W2.3：「彭奕翔網路歷程.xlsx」迄話端複合欄
    "終話基地台": "cell_id_compound",   # W2.3：同義別名，預留其他 carrier
    # 雙向通聯「基地台編號N/位置N」：cell_id 與地址用 "/" 合併，由
    # _split_compound_cell 的斜線分隔分支拆解。"基地台編號1/位置1" 為起話端、
    # "基地台編號2/位置2" 為迄話端（複合欄走 fallback 語意，前者先填即優先）。
    "基地台編號1/位置1": "cell_id_compound",
    "基地台編號2/位置2": "cell_id_compound",
    # 其他
    "細胞名稱": "sector_name",
    "小區名稱": "sector_name",
    "台號": "site_code",
    "站號": "site_code",
    "站名": "site_code",
    "細胞": "sector_id",
    "小區": "sector_id",
    "cell": "sector_id",
    "方位": "azimuth",
    "方位角": "azimuth",
    "azimuth": "azimuth",
    # ── 經緯度直給格式（GPS 軌跡 / 已含座標的檔；無 cell_id/地址，免 geocode）──
    # 例：RFL-8271軌跡.xlsx（車號 / GPS時間 / 經度 / 緯度）。
    # 注意：實務上「經度」「緯度」欄常被標反（值對調），故 _resolve_latlng 會
    # 再用「緯度必落在 [-90,90]」做範圍自動校正，不單純信任標頭。
    "GPS時間": "start_ts",
    "gps時間": "start_ts",
    "定位時間": "start_ts",
    "經度": "lng",
    "緯度": "lat",
    "經度(wgs84)": "lng",
    "緯度(wgs84)": "lat",
    "longitude": "lng",
    "latitude": "lat",
    "lng": "lng",
    "lat": "lat",
    "lon": "lng",
}
# 注意：HEADER_MAP 保留為 module-level 物件以維持向後相容
# （任何外部模組若直接 import HEADER_MAP 不會壞），
# 但實際 _normalize_row 已不使用此常數，改走 carrier_profile service
HEADER_MAP = {_canon(k): v for k, v in _RAW2CANON.items()}


# ── W2.4：方言（dialect）系統 ────────────────────────────────────────
# 為什麼需要：某些電信業者的欄位語意跟標準方言衝突，無法靠全域 alias 解。
# 例如「中華上網方言」（周蔓達上網歷程.xlsx 12869 列實測 100% 中華上網）：
#   - 起台 = ISO 8601 時間戳   ← 標準方言 W1 寫成 cell_id（錯）
#   - 起址 = cell_id 短數字     ← 標準方言 W1 寫成 cell_addr（錯）
#   - 通話對象 = 基地台地址     ← 標準方言以為是另一通話方號碼
#
# 設計取捨（為什麼不直接修 _RAW2CANON）：
#   _iter_rows_excel 的 header detection（line ~234）用 active_map 計分。
#   如果移除「起台 / 起址」全域對應，周蔓達 sheet 的真表頭命中分數會降到
#   0、整個 sheet 被跳過 → 災難。保留全域對應當「header detection 訊號」、
#   用 dialect override map 當「實際 normalize 規則」，兩者解耦。
#
# Dialect 命中時整批替換 active_map（不再走 W1 全域 alias），確保中華上網
# 方言的 row 行為由 _DIALECT_HEADER_MAPS 完全決定、可控可測。
_DIALECT_HEADER_MAPS: Dict[str, Dict[str, Optional[str]]] = {
    "cht_internet": {
        # ── 中華上網方言核心三欄（周蔓達實測）──
        "起台":     "start_ts",   # ISO 8601 時間戳，非 cell_id
        "起址":     "cell_id",    # 短數字 cell_id，非地址
        "通話對象": "cell_addr",  # 基地台地址，非另一方號碼
        # ── 方言下無意義的欄位（明確 None 表示「跳過」）──
        # 為什麼明確列出而非 implicit 跳過：dialect map 採「整批替換」，
        # 沒列在 map 內的 raw key 都會被 _normalize_row 視為未知欄而跳過，
        # 但寫出來能讓人讀 code 時清楚知道「這些欄被刻意忽略」。
        "編號":      None,   # 流水號
        "調閱號碼":  None,   # 用戶手機號（不入 record_table，PII 隔離）
        "申設人":    None,   # 申設人（PII）
        "IMEI":      None,   # 設備 ID（PII）
        "通話類別":  None,   # dialect 判斷後不需要
        "轉接電話":  None,
        "備考":      None,
        "秒數":      None,   # 上網事件無時長語意
        "始話日期":  None,   # 此 dialect 真實時間在「起台」，這欄常空
        "始話時間":  None,
        "迄台":      None,   # 此 dialect 100% 空（單端事件）
        "迄址":      None,
    },
}


def _detect_dialect(
    headers: Iterable[str],
    sample_rows: List[Dict[str, Any]],
) -> Optional[str]:
    """
    偵測 sheet 的方言。回傳 dialect key 或 None（None = 標準方言）。

    Sheet-level 而非 row-level：實測周蔓達.xlsx 12869 列 100% 都是
    「中華上網」、無混合。其他 4 個樣本的 headers 都不含「起台 + 起址」
    這對指紋欄，所以 detector 不會誤觸。

    Two-signal 驗證（避免誤觸）：
      訊號 A（必要）：headers 同時包含「起台」和「起址」
        - 單看任一欄不夠，「起台」單獨可能在其他 carrier 是 cell_id
      訊號 B（必要）：抽樣前 20 row，「通話類別」≥ 50% 含「上網」
        - 純通話 carrier 也可能用同樣 header schema、但通話類別不是上網
        - ≥ 50% 而非 100% 是因為實測同檔內可能有極少數空值或別類別

    為什麼選 20 row 為樣本：足以避開頭部 metadata 雜訊，又不會掃太多
    傷效能；實測周蔓達 row 0 起就有「中華上網」資料。
    """
    h = {str(x).strip() for x in headers if x is not None and str(x).strip()}
    # 訊號 A：必要 header 指紋
    if not ({"起台", "起址"} <= h):
        return None
    # 訊號 B：通話類別 ≥ 50% 含「上網」
    cats = []
    for r in sample_rows[:20]:
        v = str(r.get("通話類別") or "").strip()
        if v:
            cats.append(v)
    if not cats:
        # 沒有可判斷的「通話類別」樣本 → 保守拒絕（寧可走標準方言）
        return None
    internet_hits = sum(1 for c in cats if "上網" in c)
    if internet_hits / len(cats) < 0.5:
        return None
    return "cht_internet"


def _split_compound_cell(v: Any) -> Tuple[Optional[str], Optional[str]]:
    """
    W2.3：把「cell_id + 空格 + 地址 + 可選代次標籤」的複合欄拆成
    (cell_id, cell_addr)。源於彭奕翔網路歷程.xlsx 的「迄基地台」欄
    實測格式 100% 統一：'46601493130200051012 新北市中和區...(4G)'。

    切法：用第一段空白為分隔；前段視為 cell_id，後段（含可能的代次
    標籤）視為 cell_addr。代次標籤 `(4G)` / `(5G)` 等保留在 cell_addr
    內，符合 forensic「保留原始」原則（將來分析需要可再 regex 抽出）。

    斜線分隔擴充（雙向通聯）：
      部分業者把 cell_id 與地址用 "/"（或全形「／」）合併在同一欄，
      如「基地台編號1/位置1」欄值 `26634353/台北市中山區長春路31號9樓頂`。
      規則：以第一個 "/" 切兩段，僅在「左段不含中文（ID-like）且右段含
      中文（地址）」時採用，避免誤切地址內部本身帶 "/" 的情況（如
      樓層「3/4樓」）。不符合此特徵者落回原空白分隔邏輯。

    邊界處理：
      - 空 / None → (None, None)
      - 只有單一 token 且含中文 → 視為純地址：(None, addr)
      - 只有單一 token 且不含中文 → 視為純 ID：(cell_id, None)
      - 「     單純空白」→ (None, None)
    """
    if v is None:
        return (None, None)
    s = str(v).strip()
    if not s:
        return (None, None)
    # 斜線分隔：cell_id/地址（雙向通聯「基地台編號N/位置N」欄）
    slash = re.split(r"[/／]", s, maxsplit=1)
    if len(slash) == 2:
        left, right = slash[0].strip(), slash[1].strip()
        left_has_cjk = any('一' <= c <= '鿿' for c in left)
        right_has_cjk = any('一' <= c <= '鿿' for c in right)
        if left and right and not left_has_cjk and right_has_cjk:
            return (left, right)
    parts = s.split(None, 1)  # 任意空白切，最多切 1 次
    if len(parts) == 2:
        return (parts[0], parts[1].strip())
    # 單一 token：用是否含中文分辨「純 ID」還是「純地址」
    if any('\u4e00' <= c <= '\u9fff' for c in s):
        return (None, s)
    return (s, None)


def _normalize_row(
    r: Dict[str, Any],
    dialect: Optional[str] = None,
) -> Dict[str, Any]:
    """
    把原始 row dict（來源欄名）正規化成 canonical row dict。
    W1 起改從 carrier_profile service 取對照表（DB 為 SoT），
    DB 不可用時 service 會自動 fallback 到本檔的 _RAW2CANON。

    W1.5 修補（2026-04-28）：多源欄位空值不覆蓋 bug
    ────────────────────────────────────────────────
    問題：多個原始欄名可能映射到同一 canonical key（例如「基地台 ID」與
          「最終基地台 ID」皆 → cell_id）。dict 走訪順序若空值在後，
          舊版 `out[key] = v` 會把先前已寫入的有效值蓋成空字串，導致
          下游 `not cell_addr and not cell_id` 判斷誤殺整列。
    解法：對 None / 空字串 / 純空白值直接 continue，不回寫到 out；
          這形同「先到先得 + 非空覆蓋」的 fallback 語意，與電信業者
          多版本欄位（最終/起始/當前）的實務一致。

    W2.3 擴充（2026-04-29）：複合欄拆解
    ─────────────────────────────────────
    問題：彭奕翔網路歷程.xlsx 的「基地台」欄全空、真實資訊在「迄基地台」
          欄，且該欄是「ID + 空格 + 地址 + (4G)」的複合格式。
    解法：新 canonical key `cell_id_compound` 作為 dispatch 標記；
          走 _split_compound_cell 拆出 cell_id / cell_addr，分填到對應 key。
    與 W1.5 共存：採兩階段 normalize：
      Pass 1：所有直接欄 → 走 W1.5 既有「後者覆蓋」語意（行為完全不變）
      Pass 2：複合欄拆解結果 → 只在 cell_id / cell_addr 仍空時填入
              （fallback 角色，絕不蓋過原生欄位）
    這個設計的取捨：
      - 不破壞 W1.5 任何既有測試（直接欄路徑零異動）
      - 複合欄永遠是 fallback 而非 override（forensic「直接資料優先」紀律）

    W2.4 擴充（2026-04-29）：方言（dialect）系統
    ─────────────────────────────────────────────
    參數 dialect：optional dialect key（如 "cht_internet"）。
      - dialect=None（預設）→ 走原 W1.5/W2.3 邏輯，行為完全不變
      - dialect 命中（如 "cht_internet"）→ 整批替換 header_map 為
        _DIALECT_HEADER_MAPS[dialect]，跳過 W1.5 多源 fallback 與
        W2.3 複合欄拆解（dialect 內無複合欄需求）

    為什麼 dialect path 不繼承 W1.5/W2.3：
      - dialect 是「整批替換規則」、本來就要可預測；繼承會讓行為依賴
        全域 alias 順序，違反 dialect 隔離設計初衷
      - 中華上網方言實測無多源、無複合欄，根本不需這些 fallback 機制
    """
    # W2.4：若呼叫端未顯式傳 dialect，從 raw row 中讀取 _iter_rows_excel
    # 注入的 dialect tag。這讓既有呼叫 `_normalize_row(raw)` 完全不需改、
    # 也支援測試時顯式 `_normalize_row(raw, dialect="cht_internet")`。
    if dialect is None and r:
        tag = r.get("__celltrail_dialect__")
        if isinstance(tag, str) and tag:
            dialect = tag

    # 若 dialect 命中，走獨立路徑（整批替換 header_map）
    if dialect and dialect in _DIALECT_HEADER_MAPS:
        return _normalize_row_dialect(r, dialect)

    # ── 以下為原 W1.5 / W2.3 路徑（dialect=None 時的行為，零改動）──
    # lazy import 避免 circular（service 也可能 import ingest 做 fallback）
    from app.services.carrier_profile import get_active_header_map
    header_map = get_active_header_map()
    out: Dict[str, Any] = {}
    compound_pending: List[Tuple[str, Any]] = []  # (raw_key, value)

    # Pass 1：直接欄（W1.5 既有邏輯，行為不變）
    for k, v in (r or {}).items():
        # W2.4：跳過 dialect tag injection key（若不慎流入）
        if k == "__celltrail_dialect__":
            continue
        key = header_map.get(_canon(k))
        if not key:
            continue
        # 多源欄位 fallback：空值（None / 空字串 / 純空白）不覆蓋已有值
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        if key == "cell_id_compound":
            # 複合欄延後處理（保證直接欄都走完才填空缺）
            compound_pending.append((k, v))
            continue
        # 內容驗證：cell_addr 不接受 hex 短碼（LAC+CI 之類，如 `0E2921B7`）。
        # 背景：實測 0517test/台哥大-第二類.xlsx 有 ≥2 欄都映到 cell_addr
        #   （疑為「起址」hex + 「基地台地址」文字），1866 列真地址覆蓋 hex
        #   而 69 列真地址欄為空 → hex 殘留 → 送 geocoder 注定失敗。
        # 策略：hex 短碼改寫到 sector_id（若仍空），保留資訊不丟、cell_addr
        #   留空讓 coverage 把這列正確歸類為 cellid_only（需業者表）而非
        #   addr_geocode_failed（誤以為地址查不到）。
        if key == "cell_addr" and isinstance(v, str) \
                and re.fullmatch(r"[0-9A-Fa-f]{6,12}", v.strip()):
            if not out.get("sector_id"):
                out["sector_id"] = v.strip()
            continue
        out[key] = v

    # Pass 2：複合欄拆解（W2.3）— 只在對應 canonical key 仍空時填入
    for _k, v in compound_pending:
        cid, addr = _split_compound_cell(v)
        if cid and not (str(out.get("cell_id") or "").strip()):
            out["cell_id"] = cid
        if addr and not (str(out.get("cell_addr") or "").strip()):
            out["cell_addr"] = addr

    return out


def _normalize_row_dialect(
    r: Dict[str, Any],
    dialect: str,
) -> Dict[str, Any]:
    """
    W2.4：dialect-specific normalize path。

    用 _DIALECT_HEADER_MAPS[dialect] 整批替換 header_map，跳過 W1.5 多源
    fallback 與 W2.3 複合欄拆解。空值處理仍遵守「空值不覆蓋」（避免單一
    raw key 在不同 row 出現空字串時破壞 ingest 結果）。

    為什麼不需要兩階段（Pass 1/Pass 2）：
      dialect map 採「整批替換」、無多 raw key → 同 canonical 的 fallback
      語意需求；dialect 內也無複合欄定義（中華上網實測無此格式）。
    """
    dmap = _DIALECT_HEADER_MAPS[dialect]
    out: Dict[str, Any] = {}
    for k, v in (r or {}).items():
        if k == "__celltrail_dialect__":
            continue
        # dialect map 用原 raw key 直接 lookup（不過 _canon），因為 dialect
        # 內欄名都是已知精確字串（測試亦以精確字串斷言）。若未來方言量
        # 大、需 fuzzy match，可再改用 _canon。
        target = dmap.get(str(k).strip())
        if not target:
            # None（明確跳過）或 raw key 不在 dialect map → 跳過
            continue
        # 空值不覆蓋（與 W1.5 一致原則，避免空字串破壞已寫入的 row）
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        out[target] = v
    return out

# ====== DB 寫入（避免 prepared statement 衝突） ======
def _insert_records(records: List[Dict[str, Any]]) -> int:
    """批次寫入 raw_traces；座標自帶 geom"""
    if not records:
        return 0
    sql = """
    INSERT INTO raw_traces (
      project_id, target_id, start_ts, end_ts,
      cell_id, cell_addr, sector_name, site_code, sector_id,
      azimuth, lat, lng, accuracy_m, geom
    )
    VALUES (
      %(project_id)s::text, %(target_id)s::text,
      %(start_ts)s::timestamptz, %(end_ts)s::timestamptz,
      %(cell_id)s::text, %(cell_addr)s::text, %(sector_name)s::text, %(site_code)s::text, %(sector_id)s::text,
      %(azimuth)s::int, %(lat)s::float8, %(lng)s::float8, %(accuracy_m)s::int,
      CASE
        WHEN %(lat)s::float8 IS NOT NULL AND %(lng)s::float8 IS NOT NULL
          THEN ST_SetSRID(ST_MakePoint(%(lng)s::float8, %(lat)s::float8), 4326)
        ELSE NULL
      END
    )
    """
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                # psycopg3 預設多次執行後可能 server-side prepare；
                # 這裡用 executemany 但不手動 prepare，並仰賴 get_conn() 裡的 prepare_threshold=0（已在 session.py 設定）
                cur.executemany(sql, records)
        return len(records)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"匯入失敗：{type(e).__name__}: {e}")

# ====== 主要匯入（自動判斷副檔名；PDF 直接支援） ======
def ingest_auto(project_id: str, target_id: str, filename: str, file_bytes: bytes) -> Dict[str, Any]:
    """
    前端一律呼叫這支：依副檔名自動分流
    - .csv/.txt/.tsv → 文字表格
    - .xlsx → Excel
    - .pdf → 走 PDF 解析（手機拍來的 PDF 也可）

    加密（密碼保護）的 Office 檔會在此被擋下並回清楚錯誤（不嘗試解密）。
    """
    _reject_if_encrypted(file_bytes)
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in {"csv", "txt", "tsv"}:
        return _ingest_rows_stream(project_id, target_id, _iter_rows_csv(file_bytes))
    # xlsx / xltx（Excel 範本）/ xlsm（含巨集）/ xltm（含巨集範本）皆走同一條 openpyxl 路徑
    elif ext in {"xlsx", "xltx", "xlsm", "xltm"}:
        return _ingest_rows_stream(project_id, target_id, _iter_rows_excel(file_bytes))
    elif ext == "pdf":
        return ingest_pdf(project_id, target_id, file_bytes)
    else:
        raise ValueError("不支援的檔案格式：請使用 CSV / TXT / XLSX / XLTX / PDF")

def _ingest_chunk_size() -> int:
    """每塊處理的列數（env `INGEST_CHUNK_SIZE` 可調，預設 800，合理 100~5000，非法 fallback 800）。

    為什麼分塊（見 CLAUDE.md 五-R / 七-8）：整檔 hold records + 逐筆 geocode 會在
    Render 小實例 OOM/逾時。分塊讓記憶體峰值限於一塊，且每塊走 geocode.lookup_bulk
    （並行 Google + SQL geocode_cache + 去重），不再逐筆序列。
    """
    raw = os.getenv("INGEST_CHUNK_SIZE", "800")
    try:
        n = int(raw)
    except (TypeError, ValueError):
        return 800
    return n if 100 <= n <= 5000 else 800


def _ingest_rows_stream(project_id: str, target_id: str, rows_iter: Iterable[Dict[str, Any]]) -> Dict[str, Any]:
    """chunk-based 串流匯入（P8.1，2026-06-28）：每塊 normalize → bulk geocode → insert → 釋放。

    與舊版（整檔 hold to_insert + 逐筆 geocode.lookup）對比：
      - 記憶體峰值限於一塊（INGEST_CHUNK_SIZE，預設 800），不再整檔累積 → 解 OOM（七-8）。
      - geocode 每塊走 lookup_bulk（並行 + SQL 快取 + 去重）→ 解 H1（存檔路徑原本沒並行）。
      - 統計 total/inserted/skipped 語意與舊版一致；errors 仍保留 [:50] 上限。
      - 原子性「方案 A」：每塊 _insert_records 後即 commit（autocommit）。某塊寫入失敗即
        **停止後續、誠實回報「部分匯入」**（errors 首列標明 + inserted < total），由上層
        以真實 inserted/skipped 回填 evidence_stats/audit，不假裝整檔成功。

    刻意不動 _normalize_row / _parse_ts / _resolve_latlng / _guess_accuracy 的語意。
    """
    chunk_size = _ingest_chunk_size()
    total = inserted = skipped = 0
    errors: List[str] = []
    # 一塊內待 geocode/寫入的暫存；每塊 flush 後清空（記憶體不累積整檔）。
    # 每筆多帶一個 reserved key `_geo_key`：(cell_id, cell_addr) 或 None（已有直給座標）。
    pending: List[Dict[str, Any]] = []

    def _flush() -> bool:
        """處理 pending 一塊：bulk geocode → 填座標 → insert。
        回傳 True=成功；False=寫入失敗（部分匯入，呼叫端應停止後續）。"""
        nonlocal inserted
        if not pending:
            return True
        # 收集本塊需 geocode 的 unique (cell_id, cell_addr)；
        # 直給座標者（_geo_key=None）不查 → 同址在同塊內只進 lookup_bulk 一次（set 去重）。
        keys = list({p["_geo_key"] for p in pending if p["_geo_key"] is not None})
        bulk = geocode.lookup_bulk(keys) if keys else {}
        records: List[Dict[str, Any]] = []
        for p in pending:
            gk = p.pop("_geo_key")
            lat, lng = p["lat"], p["lng"]
            if gk is not None:
                ll = bulk.get(gk)
                if ll:
                    lat, lng = ll
            p["lat"] = _to_float(lat)
            p["lng"] = _to_float(lng)
            records.append(p)
        try:
            inserted += _insert_records(records)
        except HTTPException as e:
            # 方案 A：誠實回報部分匯入。插在 errors 首列以 survive [:50] 截斷。
            detail = getattr(e, "detail", str(e))
            errors.insert(0, (
                f"⚠ 部分匯入（未完全成功）：寫入資料庫時失敗，已成功寫入 {inserted} 筆、"
                f"其餘未寫入（原因：{detail}）。請排除問題後，建議用新案件重新上傳整檔，"
                f"避免與已寫入資料重複。"
            ))
            pending.clear()
            return False
        pending.clear()
        return True

    aborted = False
    for idx, raw in enumerate(rows_iter, start=1):
        total += 1
        r = _normalize_row(raw)

        start_ts = _parse_ts(r.get("start_ts"))
        end_ts = _parse_ts(r.get("end_ts")) or start_ts
        if not start_ts:
            skipped += 1
            errors.append(f"row{idx}: 缺少開始連線時間")
            continue

        cell_id = (str(r.get("cell_id") or "").strip() or None)
        cell_addr = (str(r.get("cell_addr") or "").strip() or None)
        sector_name = (str(r.get("sector_name") or "").strip() or None)
        site_code = (str(r.get("site_code") or "").strip() or None)
        sector_id = (str(r.get("sector_id") or "").strip() or None)
        azimuth = _to_int(r.get("azimuth"))

        # 檔案自帶座標（GPS 軌跡 / 經緯度格式）優先；無則走 cell_id/addr geocode
        direct_ll = _resolve_latlng(r)

        if not cell_addr and not cell_id and not direct_ll:
            skipped += 1
            errors.append(f"row{idx}: 地址、cell_id、經緯度皆空，無法定位")
            continue

        lat = lng = None
        accuracy_m = _guess_accuracy(cell_addr)
        geo_key: Optional[tuple] = None
        if direct_ll:
            # 檔案已含座標 → 直接採用，免 geocode（GPS 定位精度高）
            lat, lng = direct_ll
            accuracy_m = GPS_ACCURACY_M
        else:
            # 延後到 _flush 用 lookup_bulk 一次解析整塊（含 SQL 快取 + 並行）
            geo_key = (cell_id, cell_addr)

        pending.append(
            dict(
                project_id=project_id,
                target_id=target_id,
                start_ts=start_ts,
                end_ts=end_ts,
                cell_id=cell_id,
                cell_addr=cell_addr,
                sector_name=sector_name,
                site_code=site_code,
                sector_id=sector_id,
                azimuth=azimuth,
                lat=lat,
                lng=lng,
                accuracy_m=_to_int(accuracy_m),
                _geo_key=geo_key,
            )
        )

        if len(pending) >= chunk_size:
            if not _flush():
                aborted = True
                break

    if not aborted:
        _flush()  # 收尾剩餘（若此塊也失敗，_flush 內已記 errors）

    return {"total": total, "inserted": inserted, "skipped": skipped, "errors": errors[:50]}

# ====== PDF 匯入（不需先轉檔） ======
# pdfplumber 為可選依賴，缺少時給清楚訊息
try:
    import pdfplumber  # type: ignore
except Exception:  # pragma: no cover
    pdfplumber = None

def _match_col_idx(headers: List[str]) -> Dict[str, int]:
    """以標準化表頭來找出欄位位置；找不到回傳 -1。

    W2.5：兩階段比對，避免子字串歧義（如「細胞名稱」與「細胞」並存時的衝突）：
      Pass 1（精確）：每個 cands key 找 canon-equal 的 header；命中則認領該 index。
      Pass 2（子字串備援）：未命中的 cands key 用「`c in name`」鬆散比對，
                            但跳過 Pass 1 已認領的 index，確保「細胞」不會誤抓到
                            「細胞名稱」上。同時保留向後相容（如「基地臺編號」
                            的「臺/台」異體字仍可由 Pass 2 命中）。
    """
    h = [_canon(x) for x in headers]
    cands = {
        "start": [_canon(x) for x in ["開始連線時間", "開始時間", "起始時間", "GPS時間", "定位時間", "進入基地台時間"]],
        "end":   [_canon(x) for x in ["結束連線時間", "結束時間", "終止時間", "離開基地台時間"]],
        "cellid":[_canon(x) for x in ["基地台編號", "基地臺編號", "站台編號", "站碼", "cell_id"]],
        "addr":  [_canon(x) for x in ["基地台地址", "基地臺地址", "站台地址", "地址"]],
        "sector":[_canon(x) for x in ["細胞名稱", "小區名稱"]],
        "site":  [_canon(x) for x in ["台號", "站號", "站名"]],
        "cid":   [_canon(x) for x in ["細胞", "小區", "cell"]],
        "az":    [_canon(x) for x in ["方位", "方位角", "azimuth"]],
        # GPS 軌跡 / 經緯度直給格式（PDF 版，如 RFX-6179.pdf）
        "lat":   [_canon(x) for x in ["緯度", "latitude", "lat"]],
        "lng":   [_canon(x) for x in ["經度", "longitude", "lng", "lon"]],
    }
    got: Dict[str, int] = {key: -1 for key in cands}
    claimed: set = set()

    # Pass 1：精確匹配（canon equal）優先，避免子字串歧義
    for key, cs in cands.items():
        for i, name in enumerate(h):
            if i in claimed:
                continue
            if name in cs:
                got[key] = i
                claimed.add(i)
                break

    # Pass 2：未命中的 key 用子字串比對作備援，但跳過已認領 index
    for key, cs in cands.items():
        if got[key] >= 0:
            continue
        for i, name in enumerate(h):
            if i in claimed:
                continue
            if any(c in name for c in cs):
                got[key] = i
                claimed.add(i)
                break

    return got

def _pdf_cols_useful(col: Dict[str, int]) -> bool:
    """
    判斷 _match_col_idx 的結果是否含「至少一個可用欄位」，用來辨別某頁表格的
    第一列究竟是真表頭、還是資料列。

    為什麼需要：部分多頁 PDF 的表頭只印在第一頁，後續頁直接接資料。原本
    每頁都把第一列當表頭去比對 → 後續頁拿資料列當表頭、全數比對失敗被丟。
    遇到「第一列比不出任何欄位」時，呼叫端會沿用上一頁的表頭/欄位對應、
    並把這一列也視為資料（見 ingest_pdf / _parse_pdf_to_records）。
    """
    return any(col.get(k, -1) >= 0 for k in ("start", "cellid", "addr", "lat", "lng"))

def _split_columns_fallback(lines: List[str]) -> List[List[str]]:
    """無表格線時，以多空白或 tab 嘗試切欄"""
    splitter = r"(?:\s{2,}|\t+)"
    rows = []
    for ln in lines:
        cols = [s for s in re.split(splitter, ln) if s.strip()]
        rows.append(cols)
    return rows

def _extract_tables_from_page(page) -> List[List[List[str]]]:
    """先用表格線策略，失敗再文字行距拆欄"""
    tables: List[List[List[str]]] = []
    try:
        tables = page.extract_tables(
            {
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance": 3,
                "intersection_x_tolerance": 5,
                "intersection_y_tolerance": 5,
            }
        ) or []
    except Exception:
        tables = []
    if not tables:
        try:
            text = page.extract_text() or ""
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            if lines:
                # 找到第一行表頭（含任一已知欄名）
                header_idx = -1
                for i, ln in enumerate(lines):
                    canon_ln = _canon(ln)
                    if any(_canon(k) in canon_ln for k in _RAW2CANON.keys()):
                        header_idx = i
                        break
                if header_idx >= 0:
                    hdr = _split_columns_fallback([lines[header_idx]])[0]
                    body = _split_columns_fallback(lines[header_idx + 1 :])
                    tables = [[hdr, *body]]
        except Exception:
            pass
    return tables

def ingest_pdf(project_id: str, target_id: str, file_bytes: bytes) -> Dict[str, Any]:
    if pdfplumber is None:
        raise HTTPException(status_code=400, detail="後端缺少 pdfplumber：請安裝 pip install pdfplumber")
    total = inserted = skipped = 0
    errors: List[str] = []
    rows: List[Dict[str, Any]] = []

    last_header: Optional[List[str]] = None
    last_col: Optional[Dict[str, int]] = None
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for pno, page in enumerate(pdf.pages, start=1):
            tables = _extract_tables_from_page(page)
            for t in tables:
                if not t:
                    continue
                header0 = [(c or "").strip() for c in t[0]]
                col0 = _match_col_idx(header0)
                if _pdf_cols_useful(col0):
                    header, col = header0, col0
                    last_header, last_col = header, col
                    data_rows = t[1:]
                elif last_col is not None:
                    # 此頁無表頭（接續上頁，表頭只印在首頁）→ 沿用上頁對應，t[0] 也是資料
                    header, col = last_header, last_col
                    data_rows = t
                else:
                    continue  # 尚未遇到任何表頭，無法解析此表
                for ridx, row in enumerate(data_rows, start=2):
                    total += 1
                    rr = [(row[i] if i < len(row) else "") for i in range(len(header))]

                    start_ts = _parse_ts(rr[col["start"]]) if col.get("start", -1) >= 0 else None
                    end_ts = _parse_ts(rr[col["end"]]) if col.get("end", -1) >= 0 else start_ts
                    cell_id = (rr[col["cellid"]].strip() if col.get("cellid", -1) >= 0 and rr[col["cellid"]] else None)
                    cell_addr = (rr[col["addr"]].strip()   if col.get("addr", -1)   >= 0 and rr[col["addr"]]   else None)
                    sector_name = (
                        rr[col["sector"]].strip() if col.get("sector", -1) >= 0 and rr[col["sector"]] else None
                    )
                    site_code = (rr[col["site"]].strip()   if col.get("site", -1)   >= 0 and rr[col["site"]]   else None)
                    sector_id = (rr[col["cid"]].strip()    if col.get("cid", -1)    >= 0 and rr[col["cid"]]    else None)
                    azimuth = _to_int(rr[col["az"]]) if col.get("az", -1) >= 0 else None

                    # GPS 軌跡 / 經緯度直給格式（PDF 版）：取檔案自帶座標
                    direct_ll = None
                    if col.get("lat", -1) >= 0 and col.get("lng", -1) >= 0:
                        direct_ll = _resolve_latlng({
                            "lat": rr[col["lat"]], "lng": rr[col["lng"]],
                        })

                    if not start_ts:
                        skipped += 1
                        errors.append(f"page{pno} row{ridx}: 缺少開始連線時間")
                        continue
                    if not cell_addr and not cell_id and not direct_ll:
                        skipped += 1
                        errors.append(f"page{pno} row{ridx}: 地址、cell_id、經緯度皆空，無法定位")
                        continue

                    lat = lng = None
                    accuracy_m = _guess_accuracy(cell_addr)
                    if direct_ll:
                        lat, lng = direct_ll
                        accuracy_m = GPS_ACCURACY_M
                    else:
                        try:
                            ll = geocode.lookup(cell_id, cell_addr)
                            if ll:
                                lat, lng = ll
                        except Exception as e:
                            errors.append(f"page{pno} row{ridx}: geocode 失敗：{e}")

                    rows.append(
                        dict(
                            project_id=project_id,
                            target_id=target_id,
                            start_ts=start_ts,
                            end_ts=end_ts,
                            cell_id=cell_id,
                            cell_addr=cell_addr,
                            sector_name=sector_name,
                            site_code=site_code,
                            sector_id=sector_id,
                            azimuth=azimuth,
                            lat=_to_float(lat),
                            lng=_to_float(lng),
                            accuracy_m=_to_int(accuracy_m),
                        )
                    )

    if rows:
        inserted = _insert_records(rows)

    return {"total": total, "inserted": inserted, "skipped": skipped, "errors": errors[:50]}


# ====== Parse-only（臨時模式）：解析 + geocode，不寫 DB ======

# 「手動欄位對應」用：前端標準名 → ingest 內部已知 alias（讓 _normalize_row 自然處理）
_SYSTEM_TO_ALIAS = {
    "time":    "開始時間",    # ingest 內部 _RAW2CANON 認識
    "cell_id": "基地台編號",
    "addr":    "基地台地址",
    "lat":     "緯度",
    "lng":     "經度",
}


def _read_xlsx_top_rows(file_bytes: bytes, max_rows: int) -> List[tuple]:
    """讀 active sheet 前 max_rows 列（values_only），含「假 dimension」防呆。

    為何不直接 load_workbook(read_only=True)：部分電信匯出工具產生的 xlsx 會把
    worksheet 的 <dimension> 標籤寫死成 'A1'（宣告整表只有一格）。openpyxl 在
    read_only 模式為了省記憶體會「信任」這個邊界 → 只讀到 1 列、真資料全被跳過
    （實測台哥大上網歷程 test2.xlsx：read_only 只回 1 列，read_only=False 回 21785 列）。

    對策：先走 read_only 快路徑；若結果「退化」就 fallback 用 read_only=False 重開
    （openpyxl 非 read_only 會自行重掃真實使用範圍，不信任 dimension）。

    退化偵測（為何不是 len(rows)<=1）：對假 dimension 檔給定 max_row=N 時，
    read_only 會回「N 列但每列幾乎全空」（信任 A1 邊界，A1 以外都讀成空），
    不是只回 1 列。故改判「最寬的一列也只有 ≤1 個非空格」才是真退化徵狀
    （實測 test2.xlsx：read_only 35 列最寬僅 1 格；read_only=False 則正常 10 欄）。

    安全/成本：這兩個 peek 只在「解析失敗 → 診斷／手動對應」流程被呼叫（量小）；
    正常檔 read_only 就會回多欄、不觸發 fallback，故全載入成本只在退化檔上付。
    """
    from openpyxl import load_workbook

    def _grab(read_only: bool) -> List[tuple]:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=read_only, data_only=True)
        try:
            ws = wb.active
            return list(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True))
        finally:
            wb.close()

    def _degenerate(rows: List[tuple]) -> bool:
        widest = 0
        for r in rows:
            w = sum(1 for c in r if c is not None and str(c).strip() != "")
            if w > widest:
                widest = w
                if widest > 1:
                    return False
        return widest <= 1

    rows = _grab(True)
    if _degenerate(rows):
        rows = _grab(False)
    return rows


def _guess_header_row_idx(rows: List[tuple]) -> int:
    """從前幾列「結構性」猜真表頭在第幾列（給 peek / 手動對應用，不靠別名）。

    為什麼需要：手動對應的前提是「系統不認識此格式」，此時無法用 canonical
    別名定位表頭（_iter_rows_excel 那套會 0 命中）。但電信歷程檔的版面有穩定
    結構特徵可利用：真表頭前堆的查詢條件 / 個資區塊多半每列只有 1～3 個非空格
    （key:value 形式），真表頭與其下的資料列則「較寬且寬度一致」。

    演算法：取最寬列的寬度 max_w，門檻 thr=max(3, ⌈max_w/2⌉)；回傳第一個
    「自己 ≥thr 且下一列也 ≥thr」的列索引（= 表頭，後面緊接資料）。都不滿足
    就退回最寬的單列；空輸入回 0。

    侷限（誠實標註）：此為啟發式，非保證正確；故手動對應 UI 仍同時秀「範例值」
    讓使用者用眼睛確認哪欄是時間/地點，不單靠這個猜測。
    """
    if not rows:
        return 0

    def _width(r) -> int:
        return sum(1 for c in r if c is not None and str(c).strip() != "")

    widths = [_width(r) for r in rows]
    max_w = max(widths)
    if max_w <= 0:
        return 0
    thr = max(3, (max_w + 1) // 2)
    for i in range(len(rows) - 1):
        if widths[i] >= thr and widths[i + 1] >= thr:
            return i
    return max(range(len(rows)), key=lambda i: widths[i])


def _peek_headers(filename: str, file_bytes: bytes) -> List[str]:
    """
    抓取檔案表頭，供診斷與「手動對應」介面顯示欄位清單。

    2026-06-27：改用 _guess_header_row_idx 結構性定位真表頭，不再死抓第一個
    物理列。根因：台哥大上網歷程等格式真表頭埋在第 27 列（前面是大標題 + 查詢
    條件 + 個資區塊），舊版抓第一列只會回大標題「台灣大哥大…查詢」一個假欄位，
    害手動對應 modal 顯示錯的可對應欄位、根本無法操作。
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    try:
        if ext in {"csv", "txt", "tsv"}:
            txt = file_bytes.decode("utf-8-sig", errors="replace")
            first = txt.splitlines()[0] if txt.splitlines() else ""
            # 試逗號→tab→分號順序
            for delim in [",", "\t", ";"]:
                if delim in first:
                    return [c.strip() for c in first.split(delim)]
            return [first.strip()] if first.strip() else []
        elif ext in {"xlsx", "xltx", "xlsm", "xltm"}:
            # 多讀幾列才能涵蓋埋深表頭（對齊 _iter_rows_excel SCAN_WINDOW=30 + buffer）
            rows = _read_xlsx_top_rows(file_bytes, 35)
            if not rows:
                return []
            h = _guess_header_row_idx(rows)
            header_row = rows[h]
            return [str(c).strip() if c is not None else "" for c in header_row]
        elif ext == "pdf":
            if pdfplumber is None:
                return []
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    tables = _extract_tables_from_page(page)
                    for t in tables:
                        if t and t[0]:
                            return [(str(c) or "").strip() for c in t[0]]
                    break
            return []
    except Exception as e:
        print(f"[peek_headers] failed: {type(e).__name__}: {e}")
    return []


def _peek_sample_rows(filename: str, file_bytes: bytes, n: int = 3) -> List[List[str]]:
    """
    抓「表頭之後」的前 n 列資料，供手動對應 UI 顯示範例值。

    為什麼需要：欄名各家業者用語不一、甚至空白；讓使用者靠「看欄位內容」
    （如看到 2026-01-01 12:00 → 這是時間欄）而非「看欄名」來指認時間/地點，
    比硬猜欄名穩健得多。每格截斷 40 字避免過長。
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    out: List[List[str]] = []
    clip = lambda v: ("" if v is None else str(v)).strip()[:40]
    try:
        if ext in {"csv", "txt", "tsv"}:
            lines = [ln for ln in file_bytes.decode("utf-8-sig", errors="replace").splitlines() if ln.strip()]
            delim = next((d for d in [",", "\t", ";"] if lines and d in lines[0]), ",")
            for ln in lines[1:1 + n]:
                out.append([clip(c) for c in ln.split(delim)])
        elif ext in {"xlsx", "xltx", "xlsm", "xltm"}:
            # 與 _peek_headers 一致：先結構性定位真表頭，取其「下方」n 列當範例值，
            # 否則埋深表頭檔會把個資/查詢條件列當成範例，手動對應更難猜。
            rows = _read_xlsx_top_rows(file_bytes, 35 + n)
            h = _guess_header_row_idx(rows)
            for row in rows[h + 1: h + 1 + n]:
                out.append([clip(c) for c in row])
        elif ext == "pdf":
            if pdfplumber is not None:
                with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                    for page in pdf.pages:
                        for t in _extract_tables_from_page(page):
                            if t and len(t) > 1:
                                for r in t[1:1 + n]:
                                    out.append([clip(c) for c in r])
                                return out
                        break
    except Exception as e:
        print(f"[peek_sample_rows] failed: {type(e).__name__}: {e}")
    return out


def _build_diagnosis(headers: List[str], sample_rows: Optional[List[List[str]]] = None) -> Dict[str, Any]:
    """根據 _match_col_idx 結果產出前端可讀的診斷。"""
    if not headers:
        return {
            "found_time_col": False,
            "found_time_col_name": None,
            "found_cell_id_col": False,
            "found_cell_id_col_name": None,
            "found_addr_col": False,
            "found_addr_col_name": None,
            "available_columns": [],
            "sample_rows": sample_rows or [],
        }
    col = _match_col_idx(headers)
    def name_at(i): return headers[i] if 0 <= i < len(headers) else None
    cell_idx = col.get("cellid", -1) if col.get("cellid", -1) >= 0 else col.get("cid", -1)
    return {
        "found_time_col":          col.get("start", -1) >= 0,
        "found_time_col_name":     name_at(col.get("start", -1)),
        "found_cell_id_col":       cell_idx >= 0,
        "found_cell_id_col_name":  name_at(cell_idx),
        "found_addr_col":          col.get("addr", -1) >= 0,
        "found_addr_col_name":     name_at(col.get("addr", -1)),
        "available_columns":       headers,
        # 前幾列範例值（對齊 available_columns），供「問哪欄是時間/地點」UI 顯示
        "sample_rows":             sample_rows or [],
    }


def _apply_user_mapping(rows_iter: Iterable[Dict[str, Any]], mapping: Dict[str, str]):
    """
    把使用者指定的「欄位名→系統欄位」對應套到每筆 row。
    將 raw key rename 為 _RAW2CANON 已知的 alias，後續 _normalize_row 會自然處理。
    """
    for row in rows_iter:
        new_row: Dict[str, Any] = {}
        for k, v in row.items():
            target = mapping.get(k)
            if target == "ignore":
                continue
            alias = _SYSTEM_TO_ALIAS.get(target) if target else None
            if alias:
                new_row[alias] = v
            else:
                # 未指定的欄位保留原 key（_normalize_row 自己處理 _RAW2CANON）
                new_row[k] = v
        yield new_row


def parse_file_only(
    target_id: str,
    filename: str,
    file_bytes: bytes,
    mapping: Optional[Dict[str, str]] = None,
) -> List[Dict[str, Any]]:
    """
    解析 + geocode，但不寫入 DB。供前端臨時模式使用。
    回傳 record dict list（含 lat/lng/start_ts 等所有欄位）。

    `mapping`：使用者「手動欄位對應」傳入時，格式 {raw_column_name: system_field}
              其中 system_field ∈ {'time','cell_id','addr','lat','lng','ignore'}
              system_field 會在 ingest 內部 rename 為 _RAW2CANON 認識的 alias。

    解析後若 records 為空 → raise ParseDiagnosisError（供前端展示診斷 + 回報入口）。
    加密（密碼保護）檔會先被 _reject_if_encrypted 擋下並回清楚錯誤。
    """
    _reject_if_encrypted(file_bytes)
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    # 先取得 rows iterator（PDF 走另一條：mapping 對 PDF 不適用，因為它有專用 table 流程）
    if ext in {"csv", "txt", "tsv"}:
        rows = _iter_rows_csv(file_bytes)
        if mapping:
            rows = _apply_user_mapping(rows, mapping)
        records = _parse_rows_to_records(target_id, rows)
    elif ext in {"xlsx", "xltx", "xlsm", "xltm"}:
        # user_mapping 同時傳入 _iter_rows_excel：讓使用者指定的欄位在 header
        # detection 規則 B 被算成命中，陌生格式的 sheet 才不會在 rename 前被丟棄。
        rows = _iter_rows_excel(file_bytes, user_mapping=mapping)
        if mapping:
            rows = _apply_user_mapping(rows, mapping)
        records = _parse_rows_to_records(target_id, rows)
    elif ext == "pdf":
        if mapping:
            # PDF 暫不支援使用者手動對應（內部用 _match_col_idx 直接讀 table）
            raise ValueError("PDF 暫不支援手動欄位對應，請改用 CSV/Excel")
        records = _parse_pdf_to_records(target_id, file_bytes)
    else:
        raise ValueError("不支援的檔案格式：請使用 CSV / TXT / XLSX / XLTX / PDF")

    # 解析後 0 筆 → 觸發智慧診斷（附前幾列範例值供「問哪欄是時間/地點」UI）
    if not records:
        headers = _peek_headers(filename, file_bytes)
        sample_rows = _peek_sample_rows(filename, file_bytes)
        raise ParseDiagnosisError(
            "無法從此檔案解析出任何有效記錄",
            _build_diagnosis(headers, sample_rows),
        )

    return records


def _parse_rows_to_records(target_id: str, rows_iter: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    把 rows 解析 + geocode，不做 DB insert，回傳 list[dict]。

    B 方案（批次 geocode）：
      Phase 1：掃過所有 rows → normalize + 時間驗證 → 收集 (cell_id, cell_addr)
      Phase 2：把 unique (cell_id, cell_addr) 一次丟給 geocode.lookup_bulk
               （內部一次 SQL ANY + Redis MGET，避免 3000+ round-trip）
      Phase 3：第二輪用 bulk 結果組裝 records
    """
    import time as _time
    _t_start = _time.perf_counter()
    _rows_read = 0

    # ── Phase 1：normalize + 時間驗證 + 收集 unique geo keys ────────
    _t_phase1 = _time.perf_counter()
    parsed: List[Dict[str, Any]] = []   # 暫存解析後且通過篩選的 row
    for raw in rows_iter:
        _rows_read += 1
        r = _normalize_row(raw)

        start_ts = _parse_ts(r.get("start_ts"))
        if not start_ts:
            continue
        end_ts = _parse_ts(r.get("end_ts")) or start_ts

        cell_id   = (str(r.get("cell_id")   or "").strip() or None)
        cell_addr = (str(r.get("cell_addr") or "").strip() or None)
        # 檔案自帶座標（GPS 軌跡 / 經緯度格式）優先
        direct_ll = _resolve_latlng(r)
        if not cell_addr and not cell_id and not direct_ll:
            continue

        parsed.append({
            "r":         r,
            "start_ts":  start_ts,
            "end_ts":    end_ts,
            "cell_id":   cell_id,
            "cell_addr": cell_addr,
            "direct_ll": direct_ll,
        })
    _t_phase1_elapsed = _time.perf_counter() - _t_phase1

    # ── Phase 2：批次 geocode（只查無自帶座標的列）────────────────
    _t_phase2 = _time.perf_counter()
    unique_keys = list({(p["cell_id"], p["cell_addr"]) for p in parsed if not p["direct_ll"]})
    try:
        bulk = geocode.lookup_bulk(unique_keys)
    except Exception as e:
        print(f"[ingest] bulk geocode error, fallback to None for all: {type(e).__name__}: {e}")
        bulk = {k: None for k in unique_keys}
    _t_phase2_elapsed = _time.perf_counter() - _t_phase2

    # ── Phase 3：組裝 records ──────────────────────────────────────
    _t_phase3 = _time.perf_counter()
    result: List[Dict[str, Any]] = []
    _n_geo_ok = 0
    for p in parsed:
        lat = lng = None
        accuracy_m = _guess_accuracy(p["cell_addr"])
        if p["direct_ll"]:
            # 檔案已含座標 → 直接採用，免 geocode（GPS 精度高）
            lat, lng = p["direct_ll"]
            accuracy_m = GPS_ACCURACY_M
            _n_geo_ok += 1
        else:
            ll = bulk.get((p["cell_id"], p["cell_addr"]))
            if ll:
                lat, lng = ll
                _n_geo_ok += 1
        r = p["r"]
        result.append({
            "target_id":   target_id,
            "start_ts":    p["start_ts"].isoformat(),
            "end_ts":      p["end_ts"].isoformat(),
            "cell_id":     p["cell_id"],
            "cell_addr":   p["cell_addr"],
            "sector_name": (str(r.get("sector_name") or "").strip() or None),
            "site_code":   (str(r.get("site_code")   or "").strip() or None),
            "sector_id":   (str(r.get("sector_id")   or "").strip() or None),
            "azimuth":     _to_int(r.get("azimuth")),
            "lat":         _to_float(lat),
            "lng":         _to_float(lng),
            "accuracy_m":  _to_int(accuracy_m),
            "azimuth_ref": "unknown",
        })
    _t_phase3_elapsed = _time.perf_counter() - _t_phase3

    _total = _time.perf_counter() - _t_start
    print(
        f"[ingest][timing] _parse_rows_to_records: "
        f"total={_total*1000:.0f}ms "
        f"rows_read={_rows_read} kept={len(result)} ok={_n_geo_ok} "
        f"unique_keys={len(unique_keys)} "
        f"phase1_normalize={_t_phase1_elapsed*1000:.0f}ms "
        f"phase2_bulk_geocode={_t_phase2_elapsed*1000:.0f}ms "
        f"phase3_assemble={_t_phase3_elapsed*1000:.0f}ms"
    )
    return result


def _parse_pdf_to_records(target_id: str, file_bytes: bytes) -> List[Dict[str, Any]]:
    """
    同 ingest_pdf 但不寫 DB，回傳 list[dict]。
    B 方案（批次 geocode）：先收集 unique (cell_id, cell_addr) → 一次 bulk → 再組裝。
    """
    if pdfplumber is None:
        raise ValueError("後端缺少 pdfplumber")
    import time as _time
    _t_start = _time.perf_counter()
    _t_pdf_open = 0.0
    _t_extract  = 0.0
    _rows_read   = 0

    # ── Phase 1：解析 PDF 表格 + 時間驗證 ──────────────────────────
    _t_phase1 = _time.perf_counter()
    parsed: List[Dict[str, Any]] = []

    _t0 = _time.perf_counter()
    pdf = pdfplumber.open(io.BytesIO(file_bytes))
    _t_pdf_open = _time.perf_counter() - _t0
    last_header: Optional[List[str]] = None
    last_col: Optional[Dict[str, int]] = None
    with pdf:
        for pno, page in enumerate(pdf.pages, start=1):
            _t1 = _time.perf_counter()
            tables = _extract_tables_from_page(page)
            _t_extract += _time.perf_counter() - _t1
            for t in tables:
                if not t:
                    continue
                header0 = [(c or "").strip() for c in t[0]]
                col0 = _match_col_idx(header0)
                if _pdf_cols_useful(col0):
                    header, col = header0, col0
                    last_header, last_col = header, col
                    data_rows = t[1:]
                elif last_col is not None:
                    # 此頁無表頭（接續上頁）→ 沿用上頁對應，t[0] 也是資料
                    header, col = last_header, last_col
                    data_rows = t
                else:
                    continue
                for row in data_rows:
                    _rows_read += 1
                    rr = [(row[i] if i < len(row) else "") for i in range(len(header))]
                    start_ts = _parse_ts(rr[col["start"]]) if col.get("start", -1) >= 0 else None
                    if not start_ts:
                        continue
                    end_ts = _parse_ts(rr[col["end"]]) if col.get("end", -1) >= 0 else start_ts
                    cell_id   = (rr[col["cellid"]].strip() if col.get("cellid", -1) >= 0 and rr[col["cellid"]] else None)
                    cell_addr = (rr[col["addr"]].strip()   if col.get("addr", -1)   >= 0 and rr[col["addr"]]   else None)
                    # GPS 軌跡 / 經緯度直給格式（PDF 版）：取檔案自帶座標
                    direct_ll = None
                    if col.get("lat", -1) >= 0 and col.get("lng", -1) >= 0:
                        direct_ll = _resolve_latlng({
                            "lat": rr[col["lat"]], "lng": rr[col["lng"]],
                        })
                    if not cell_addr and not cell_id and not direct_ll:
                        continue
                    parsed.append({
                        "start_ts":    start_ts,
                        "end_ts":      end_ts,
                        "cell_id":     cell_id,
                        "cell_addr":   cell_addr,
                        "direct_ll":   direct_ll,
                        "sector_name": (rr[col["sector"]].strip() if col.get("sector", -1) >= 0 and rr[col["sector"]] else None),
                        "site_code":   (rr[col["site"]].strip()   if col.get("site",   -1) >= 0 and rr[col["site"]]   else None),
                        "sector_id":   (rr[col["cid"]].strip()    if col.get("cid",    -1) >= 0 and rr[col["cid"]]    else None),
                        "azimuth":     (_to_int(rr[col["az"]]) if col.get("az", -1) >= 0 else None),
                    })
    _t_phase1_elapsed = _time.perf_counter() - _t_phase1

    # ── Phase 2：批次 geocode（只查無自帶座標的列）────────────────
    _t_phase2 = _time.perf_counter()
    unique_keys = list({(p["cell_id"], p["cell_addr"]) for p in parsed if not p["direct_ll"]})
    try:
        bulk = geocode.lookup_bulk(unique_keys)
    except Exception as e:
        print(f"[ingest][pdf] bulk geocode error, fallback to None: {type(e).__name__}: {e}")
        bulk = {k: None for k in unique_keys}
    _t_phase2_elapsed = _time.perf_counter() - _t_phase2

    # ── Phase 3：組裝 records ──────────────────────────────────────
    _t_phase3 = _time.perf_counter()
    result: List[Dict[str, Any]] = []
    _n_geo_ok = 0
    for p in parsed:
        lat = lng = None
        accuracy_m = _guess_accuracy(p["cell_addr"])
        if p["direct_ll"]:
            lat, lng = p["direct_ll"]
            accuracy_m = GPS_ACCURACY_M
            _n_geo_ok += 1
        else:
            ll = bulk.get((p["cell_id"], p["cell_addr"]))
            if ll:
                lat, lng = ll
                _n_geo_ok += 1
        result.append({
            "target_id":   target_id,
            "start_ts":    p["start_ts"].isoformat(),
            "end_ts":      p["end_ts"].isoformat(),
            "cell_id":     p["cell_id"],
            "cell_addr":   p["cell_addr"],
            "sector_name": p["sector_name"],
            "site_code":   p["site_code"],
            "sector_id":   p["sector_id"],
            "azimuth":     p["azimuth"],
            "lat":         _to_float(lat),
            "lng":         _to_float(lng),
            "accuracy_m":  _to_int(accuracy_m),
            "azimuth_ref": "unknown",
        })
    _t_phase3_elapsed = _time.perf_counter() - _t_phase3

    _total = _time.perf_counter() - _t_start
    print(
        f"[ingest][timing] _parse_pdf_to_records: "
        f"total={_total*1000:.0f}ms "
        f"pdf_open={_t_pdf_open*1000:.0f}ms extract_tables={_t_extract*1000:.0f}ms "
        f"rows_read={_rows_read} kept={len(result)} ok={_n_geo_ok} "
        f"unique_keys={len(unique_keys)} "
        f"phase1_parse_pdf={_t_phase1_elapsed*1000:.0f}ms "
        f"phase2_bulk_geocode={_t_phase2_elapsed*1000:.0f}ms "
        f"phase3_assemble={_t_phase3_elapsed*1000:.0f}ms"
    )
    return result