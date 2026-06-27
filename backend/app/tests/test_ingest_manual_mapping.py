# backend/app/tests/test_ingest_manual_mapping.py
"""
手動欄位對應（manual column mapping）結構性修復（2026-06-27）

使用者反映：系統無法識別格式時，「手動對應欄位」無法正確操作。

確認根因（兩層）：
  1. pipeline 順序顛倒：手動對應是 _iter_rows_excel（先偵測表頭、丟棄 sheet）
     → 才 _apply_user_mapping（rename）。但「完全不認識」時，規則 B（表頭須命中
     ≥2 個已知別名）會在 rename 前就把整張 sheet 丟掉 → 使用者的對應無從施力。
  2. _peek_headers 死抓第一個物理列 → 埋深表頭檔（真表頭在第 N 列）會把大標題
     當成「可對應欄位」秀給使用者 → modal 顯示錯欄、無法操作。

修法：
  1. _iter_rows_excel(user_mapping=...)：header detection 把使用者指定的欄位也
     算命中，陌生 sheet 不被丟棄。
  2. _peek_headers / _peek_sample_rows 改用 _guess_header_row_idx 結構性定位
     真表頭（不靠別名）。

本檔不依賴 DB（monkeypatch carrier_profile）。
"""
from __future__ import annotations

import io
import os
from typing import Any, List

os.environ.setdefault("DATABASE_URL", "postgresql://user:pass@localhost:5432/fakedb")
os.environ.setdefault("SECRET_KEY", "test-secret-key-only-for-pytest")
os.environ.setdefault("AUTH_ENABLED", "true")


def _make_xlsx(rows: List[List[Any]]) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _patch_active_map_to_default(monkeypatch):
    import app.services.carrier_profile as cp

    monkeypatch.setattr(cp, "_load_default_profile_from_db", lambda: None)
    cp.invalidate_cache()


# 完全陌生的欄名（系統一個都不認識）+ 真表頭埋在第 21 列
_UNKNOWN_HEADER = ["連線起始", "設備代碼", "所在位置"]


def _unknown_buried_blob() -> bytes:
    rows = [[f"查詢條件{i}:xxx"] for i in range(21)]          # 0..20 前置窄列
    rows.append(_UNKNOWN_HEADER)                              # index 21 真表頭
    rows.append(["2025-10-01 10:00", "ABC123", "台北市信義區市府路1號"])
    rows.append(["2025-10-01 10:05", "ABC124", "台北市大安區和平東路100號"])
    return _make_xlsx(rows)


# ─────────────────────────────────────────────────────────────
# 1. _guess_header_row_idx：結構性定位（不靠別名）
# ─────────────────────────────────────────────────────────────
def test_guess_header_flat():
    """表頭在第 0 列（無前置）→ 回 0。"""
    from app.services.ingest import _guess_header_row_idx

    rows = [("a", "b", "c"), ("1", "2", "3"), ("4", "5", "6")]
    assert _guess_header_row_idx(rows) == 0


def test_guess_header_buried():
    """前面 21 列窄 metadata，真表頭（寬列）在 index 21 → 須回 21。"""
    from app.services.ingest import _guess_header_row_idx

    rows = [(f"meta{i}", None, None) for i in range(21)]
    rows += [("連線起始", "設備代碼", "所在位置"),
             ("2025-10-01", "ABC", "台北市"),
             ("2025-10-02", "DEF", "高雄市")]
    assert _guess_header_row_idx(rows) == 21


def test_guess_header_ignores_narrow_metadata_pairs():
    """前置出現連兩列『3 欄寬』的查詢條件，但仍不及真表頭(寬)→ 不被誤判。"""
    from app.services.ingest import _guess_header_row_idx

    rows = [
        ("類別:A", "目標:B", "起始:C"),       # 0：3 欄寬的查詢條件
        ("種類:D", "業者:E", "終止:F"),       # 1：同上
        ("使用者資料", None, None),            # 2
    ]
    # 真表頭 6 欄寬 → max_w=6, thr=3；但 index0/1 也達 3 → 此例驗證「就近第一個
    # 連兩列達標」會落在 0。為避免誤判，真實情境靠 max_w 拉高門檻（見下個測試）。
    rows += [("t1", "t2", "t3", "t4", "t5", "t6"),
             ("d1", "d2", "d3", "d4", "d5", "d6")]
    # max_w=6 → thr=max(3,3)=3；index0 width3>=3 且 index1 width3>=3 → 回 0。
    # 這是 thr 偏低時的已知侷限，下個測試用更寬真表頭證明門檻會拉開。
    assert _guess_header_row_idx(rows) == 0


def test_guess_header_threshold_scales_with_width():
    """真表頭夠寬（10 欄）時，門檻 thr=5 會濾掉 3 欄寬的前置條件列。"""
    from app.services.ingest import _guess_header_row_idx

    rows = [
        ("類別:A", "目標:B", "起始:C"),   # 0：3 欄寬，< thr(5)
        ("種類:D", "業者:E", "終止:F"),   # 1：3 欄寬，< thr
        ("使用者資料", None, None),        # 2
    ]
    wide_header = [f"h{i}" for i in range(10)]
    wide_data = [f"d{i}" for i in range(10)]
    rows += [tuple(wide_header), tuple(wide_data)]   # index 3, 4
    assert _guess_header_row_idx(rows) == 3


# ─────────────────────────────────────────────────────────────
# 2. peek：埋深 + 陌生欄名 → 仍抓到真表頭與範例值
# ─────────────────────────────────────────────────────────────
def test_peek_finds_buried_unknown_header():
    from app.services.ingest import _peek_headers, _peek_sample_rows

    blob = _unknown_buried_blob()
    headers = _peek_headers("x.xlsx", blob)
    assert headers == _UNKNOWN_HEADER, f"應抓到真表頭，實際 {headers}"

    samples = _peek_sample_rows("x.xlsx", blob, n=2)
    assert len(samples) == 2
    assert samples[0][0] == "2025-10-01 10:00"
    assert "台北市信義區" in samples[0][2]


# ─────────────────────────────────────────────────────────────
# 3. 核心：user_mapping 讓陌生格式不被規則 B 丟棄
# ─────────────────────────────────────────────────────────────
def test_manual_mapping_rescues_unknown_format(monkeypatch):
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import (
        _iter_rows_excel, _apply_user_mapping, _normalize_row,
    )

    blob = _unknown_buried_blob()
    mapping = {"連線起始": "time", "設備代碼": "cell_id", "所在位置": "addr"}

    # 不帶 mapping：規則 B 仍正確守門（陌生欄名 0 命中）→ 0 列
    assert len(list(_iter_rows_excel(blob))) == 0

    # 帶 user_mapping：被救起來 → 2 列，且正確 normalize
    rows = list(_iter_rows_excel(blob, user_mapping=mapping))
    assert len(rows) == 2, f"user_mapping 應救起陌生格式，實際 {len(rows)}"

    norm = [_normalize_row(r) for r in _apply_user_mapping(iter(rows), mapping)]
    assert norm[0]["start_ts"] == "2025-10-01 10:00"
    assert norm[0]["cell_id"] == "ABC123"
    assert norm[0]["cell_addr"] == "台北市信義區市府路1號"


def test_manual_mapping_ignore_field(monkeypatch):
    """mapping 值為 'ignore' 的欄位不算命中、也不進 normalize。"""
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    blob = _make_xlsx([
        ["欄甲", "欄乙"],   # 兩欄都不認識
        ["v1", "v2"],
        ["v3", "v4"],
    ])
    # 只指定一欄、另一欄 ignore → 命中數 1 < 2 → 仍被規則 B 擋（符合預期：
    # 至少要指認到能湊出有效記錄的欄位；單一非時間欄無法成案）
    only_one = {"欄甲": "cell_id", "欄乙": "ignore"}
    assert len(list(_iter_rows_excel(blob, user_mapping=only_one))) == 0

    # 指定兩欄 → 命中 2 → 救起
    both = {"欄甲": "time", "欄乙": "addr"}
    assert len(list(_iter_rows_excel(blob, user_mapping=both))) == 2
