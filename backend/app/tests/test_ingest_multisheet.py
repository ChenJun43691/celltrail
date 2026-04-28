# backend/app/tests/test_ingest_multisheet.py
"""
W2.1 多 sheet 支援單元測試（2026-04-29）

驗證 _iter_rows_excel 在多 sheet xlsx 上的新行為：
  1. 單 sheet → 行為與舊版一致（向後相容）
  2. 多 sheet 且皆為資料 → 全部讀到
  3. 規則 A：sheet 行數 < 5 → 跳過
  4. 規則 B：sheet 欄名全不在 carrier dialect → 跳過（人資 sheet）
  5. A + B 同時觸發 → 只跳過一次（不重複處理）
  6. 假表頭偵測「每 sheet 獨立」（一個 sheet 真表頭 row 0、另一個 row 1）
  7. 邊界：0 個資料 sheet（全被過濾）→ 空迭代器，不 raise
  8. 跳過的 sheet 寫到 logging（forensic 軌跡）

不依賴 DB；用 monkeypatch 把 carrier_profile.get_active_header_map
patch 成回傳 ingest.HEADER_MAP，避免 DB 查詢污染測試。
"""
from __future__ import annotations

import io
import os
import logging
from typing import Any, List, Tuple

# 必須在 import app.* 之前設好環境變數，避免 db.session 在 import 期間崩潰
os.environ.setdefault("DATABASE_URL", "postgresql://user:pass@localhost:5432/fakedb")
os.environ.setdefault("SECRET_KEY", "test-secret-key-only-for-pytest")
os.environ.setdefault("AUTH_ENABLED", "true")


# ─────────────────────────────────────────────────────────────
# 共用 fixture：用 openpyxl 動態組 in-memory xlsx
# ─────────────────────────────────────────────────────────────
def _make_xlsx(sheets: List[Tuple[str, List[List[Any]]]]) -> bytes:
    """
    sheets = [(sheet_name, [[row1_cells], [row2_cells], ...]), ...]

    第一個 row 通常是表頭。openpyxl 預設不寫入合併儲存格，所以這份 fixture
    產生的 xlsx 表頭都在 row 0，方便驗證單純情境。
    若要驗證「假表頭」情境（真表頭在 row 1），請用 _make_xlsx_with_supertitle()。
    """
    import openpyxl
    wb = openpyxl.Workbook()
    # 移除預設的 'Sheet'，從乾淨狀態開始
    wb.remove(wb.active)
    for name, rows in sheets:
        ws = wb.create_sheet(title=name)
        for r in rows:
            ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _make_xlsx_with_supertitle(
    sheets: List[Tuple[str, List[Any], List[List[Any]], bool]]
) -> bytes:
    """
    sheets = [(sheet_name, header_row, data_rows, has_supertitle), ...]
      - has_supertitle=True：在 header_row 之前先寫一列「合併大標」風格列
        （第一格有值、其餘空）以觸發 _iter_rows_excel 的 Unnamed 偵測
      - has_supertitle=False：表頭直接放第 0 列

    這個 helper 模擬「電信業者真表頭埋在第 1 列」的情境。
    """
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, hdr, data, has_super in sheets:
        ws = wb.create_sheet(title=name)
        if has_super:
            # 「行動上網」「網路歷程」這種跨欄大標：第一格有值，其餘空
            ws.append(["行動上網"] + [None] * (len(hdr) - 1))
        ws.append(hdr)
        for r in data:
            ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _patch_active_map_to_default(monkeypatch):
    """
    讓 _iter_rows_excel 在規則 B 用本檔常數，避免測試碰 DB。
    回傳當前 active_map（即 HEADER_MAP），方便個別測試驗證。
    """
    import app.services.carrier_profile as cp
    monkeypatch.setattr(cp, "_load_default_profile_from_db", lambda: None)
    cp.invalidate_cache()


# ─────────────────────────────────────────────────────────────
# 1. 向後相容：單 sheet
# ─────────────────────────────────────────────────────────────
def test_single_sheet_backward_compat(monkeypatch):
    """單 sheet xlsx 應與舊版行為完全一致：讀到所有資料列"""
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    blob = _make_xlsx([
        ("通聯", [
            ["時間", "基地台", "通話對象"],
            ["2026-01-01 10:00:00", "CELL_A", "0912345678"],
            ["2026-01-01 10:05:00", "CELL_B", "0987654321"],
            ["2026-01-01 10:10:00", "CELL_C", "0911111111"],
            ["2026-01-01 10:15:00", "CELL_D", "0922222222"],
            ["2026-01-01 10:20:00", "CELL_E", "0933333333"],
        ]),
    ])
    rows = list(_iter_rows_excel(blob))
    assert len(rows) == 5
    assert rows[0]["時間"] == "2026-01-01 10:00:00"
    assert rows[0]["基地台"] == "CELL_A"
    assert rows[4]["基地台"] == "CELL_E"


# ─────────────────────────────────────────────────────────────
# 2. 多 sheet 全部資料：全讀
# ─────────────────────────────────────────────────────────────
def test_multi_sheet_all_data(monkeypatch):
    """3 個 sheet × 各 5 列 → 應讀到 15 列（這是 W2.1 的核心目的）"""
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    def _data_sheet(prefix: str) -> List[List[Any]]:
        return [
            ["時間", "基地台", "通話對象"],
            *[[f"2026-01-01 1{i}:00:00", f"{prefix}_{i}", f"091234567{i}"] for i in range(5)],
        ]

    blob = _make_xlsx([
        ("1月", _data_sheet("JAN")),
        ("2月", _data_sheet("FEB")),
        ("3月", _data_sheet("MAR")),
    ])
    rows = list(_iter_rows_excel(blob))
    assert len(rows) == 15

    # 驗證跨 sheet 順序：1月 → 2月 → 3月
    cells = [r["基地台"] for r in rows]
    assert cells[0] == "JAN_0"
    assert cells[5] == "FEB_0"
    assert cells[10] == "MAR_0"


# ─────────────────────────────────────────────────────────────
# 3. 規則 A：sheet 太短被跳過
# ─────────────────────────────────────────────────────────────
def test_skip_short_sheet(monkeypatch, caplog):
    """資料 sheet（10 列）+ 摘要 sheet（2 列）→ 只讀資料 sheet"""
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    blob = _make_xlsx([
        ("通聯", [
            ["時間", "基地台", "通話對象"],
            *[[f"2026-01-01 10:0{i}:00", f"CELL_{i}", "0911111111"] for i in range(10)],
        ]),
        ("摘要", [
            ["時間", "基地台"],
            ["2026-01-01", "TOTAL_5"],   # 只有 1 列資料 → len(df)=1 < 5
        ]),
    ])
    with caplog.at_level(logging.INFO, logger="app.services.ingest"):
        rows = list(_iter_rows_excel(blob))
    assert len(rows) == 10
    # 跳過資訊應寫進 log
    assert any("摘要" in m and "row<5" in m for m in caplog.messages), (
        f"未在 log 看到「摘要」被跳過的紀錄；實際 log：{caplog.messages}"
    )


# ─────────────────────────────────────────────────────────────
# 4. 規則 B：人資 sheet 被跳過
# ─────────────────────────────────────────────────────────────
def test_skip_no_header_match(monkeypatch, caplog):
    """資料 sheet + 人資 sheet（欄名「姓名/出生/身分證」全不命中）→ 只讀資料"""
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    blob = _make_xlsx([
        ("通聯", [
            ["時間", "基地台", "通話對象"],
            *[[f"2026-01-01 10:0{i}:00", f"CELL_{i}", "0911111111"] for i in range(10)],
        ]),
        ("基本人資", [
            ["姓名", "出生", "身分證", "電話"],
            ["王小明", "1985-01-01", "A123456789", "0911111111"],
            ["李大華", "1990-05-05", "B234567890", "0922222222"],
            ["張三", "1992-08-08", "C345678901", "0933333333"],
            ["李四", "1988-03-03", "D456789012", "0944444444"],
            ["趙五", "1995-11-11", "E567890123", "0955555555"],
        ]),
    ])
    with caplog.at_level(logging.INFO, logger="app.services.ingest"):
        rows = list(_iter_rows_excel(blob))
    assert len(rows) == 10
    # 確認 PII 欄位完全沒進入 ingest 路徑（forensic data minimization）
    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())
    assert "姓名" not in all_keys
    assert "身分證" not in all_keys
    assert any("基本人資" in m and "no header match" in m for m in caplog.messages)


# ─────────────────────────────────────────────────────────────
# 5. A + B 同時觸發：仍只跳過一次
# ─────────────────────────────────────────────────────────────
def test_skip_short_takes_precedence(monkeypatch, caplog):
    """
    既短又無命中的 sheet：規則 A 先觸發（短路），不會重複記錄。
    這個測試確保不會因為兩條規則都觸發而造成 log 噪音或邏輯錯誤。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    blob = _make_xlsx([
        ("通聯", [
            ["時間", "基地台", "通話對象"],
            *[[f"2026-01-01 10:0{i}:00", f"CELL_{i}", "0911111111"] for i in range(10)],
        ]),
        ("封面", [
            ["專案名稱", "承辦人"],
            ["XX分局通聯紀錄", "陳警官"],
        ]),
    ])
    with caplog.at_level(logging.INFO, logger="app.services.ingest"):
        rows = list(_iter_rows_excel(blob))
    assert len(rows) == 10
    # 規則 A 短路 → log 應只出現一次「封面」
    封面_logs = [m for m in caplog.messages if "封面" in m]
    assert len(封面_logs) == 1, f"應該只有一筆 log；實際：{封面_logs}"
    # 短路規則為 A（行<5），不該看到 "no header match"
    assert "row<5" in 封面_logs[0]
    assert "no header match" not in 封面_logs[0]


# ─────────────────────────────────────────────────────────────
# 6. 假表頭偵測「每 sheet 獨立」
# ─────────────────────────────────────────────────────────────
def test_per_sheet_fake_header_detection(monkeypatch):
    """
    一個 sheet 真表頭在 row 0（無 supertitle），
    另一個 sheet 真表頭在 row 1（有「行動上網」大標）。
    兩 sheet 都該被正確讀到 5 列資料。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    data_rows = [[f"2026-01-01 1{i}:00:00", f"CELL_{i}", "0911111111"] for i in range(5)]

    blob = _make_xlsx_with_supertitle([
        ("正常表頭", ["時間", "基地台", "通話對象"], data_rows, False),
        ("假表頭",   ["時間", "基地台", "通話對象"], data_rows, True),
    ])
    rows = list(_iter_rows_excel(blob))
    assert len(rows) == 10, (
        f"應讀到 10 列（兩 sheet 各 5 列），實際 {len(rows)}。"
        "可能原因：假表頭 sheet 的 Unnamed 偵測失敗。"
    )
    # 每列都該有「時間」欄（不該因偵測錯誤變成 'Unnamed:0'）
    for r in rows:
        assert "時間" in r, f"row 缺「時間」欄：keys={list(r.keys())}"


# ─────────────────────────────────────────────────────────────
# 7. 邊界：所有 sheet 都被過濾
# ─────────────────────────────────────────────────────────────
def test_all_sheets_filtered_returns_empty(monkeypatch, caplog):
    """所有 sheet 都觸發跳過規則 → 空迭代器，不該 raise"""
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    blob = _make_xlsx([
        ("封面", [
            ["專案名稱"],
            ["XX案"],
        ]),
        ("人資", [
            ["姓名", "身分證", "電話"],
            ["王小明", "A1", "0911"],
            ["李大華", "B2", "0922"],
            ["張三", "C3", "0933"],
            ["李四", "D4", "0944"],
            ["趙五", "E5", "0955"],
        ]),
    ])
    with caplog.at_level(logging.INFO, logger="app.services.ingest"):
        rows = list(_iter_rows_excel(blob))
    assert rows == []
    # 兩個 sheet 都該被記錄
    assert any("封面" in m for m in caplog.messages)
    assert any("人資" in m for m in caplog.messages)


# ─────────────────────────────────────────────────────────────
# 8. ingest 主流程整合：跳過 sheet 不該影響 inserted 計數
# ─────────────────────────────────────────────────────────────
def test_multi_sheet_with_normalize(monkeypatch):
    """
    驗證多 sheet 讀進來後，每列都能正確進到 _normalize_row → start_ts/cell_id。
    這個測試把 ingest pipeline 跑半套（不寫 DB），確保 W2.1 不會回歸破
    W1 / W1.5 已修好的行為。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel, _normalize_row, _parse_ts

    blob = _make_xlsx([
        ("1月", [
            ["時間", "基地台", "通話對象"],
            *[[f"2026-01-{1+i:02d} 10:00:00", f"JAN_{i}", "0911111111"] for i in range(5)],
        ]),
        ("人資", [
            ["姓名", "出生"],
            ["王", "1990"],
            ["李", "1991"],
            ["張", "1992"],
            ["陳", "1993"],
            ["林", "1994"],
        ]),
        ("2月", [
            ["時間", "基地台", "通話對象"],
            *[[f"2026-02-{1+i:02d} 10:00:00", f"FEB_{i}", "0922222222"] for i in range(5)],
        ]),
    ])
    raw_rows = list(_iter_rows_excel(blob))
    assert len(raw_rows) == 10  # 1月 5 + 2月 5，人資被擋

    ok = 0
    for raw in raw_rows:
        norm = _normalize_row(raw)
        st = _parse_ts(norm.get("start_ts"))
        cid = (str(norm.get("cell_id") or "").strip() or None)
        if st and cid:
            ok += 1
    assert ok == 10, "多 sheet 路徑下，所有列都該能正確 normalize"
