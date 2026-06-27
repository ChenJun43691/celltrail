# backend/app/tests/test_ingest_simple_time_location.py
"""
simple_time_location 極簡 Excel 格式（P8.2-fmt，2026-06-28）

A 欄=時間、B 欄=位置（地址或經緯度）、A/B 以外忽略、可能有/無表頭。
不靠檔名、靠結構與內容判斷；僅在 _iter_rows_excel 規則 B 失敗時作為 fallback。

驗證重點：
  - _parse_simple_time（民國年隔離、不動 _parse_ts）
  - _parse_latlng_text（雙浮點嚴格、順序自動校正、中文拒絕）
  - _iter_simple_time_location（有/無表頭、地址/經緯度、壞列 skip、雜欄忽略）
  - 不誤判台哥大 / 雙向通聯 / 一般電信格式（皆過規則 B、不進 fallback）

不依賴 DB（monkeypatch carrier_profile fallback 到 _RAW2CANON）。
"""
from __future__ import annotations

import io
import os
from datetime import datetime
from typing import Any, List

os.environ.setdefault("DATABASE_URL", "postgresql://user:pass@localhost:5432/fakedb")
os.environ.setdefault("SECRET_KEY", "test-secret-key-only-for-pytest")
os.environ.setdefault("AUTH_ENABLED", "true")


def _make_xlsx(rows: List[List[Any]]) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _patch_active_map_to_default(monkeypatch):
    import app.services.carrier_profile as cp

    monkeypatch.setattr(cp, "_load_default_profile_from_db", lambda: None)
    cp.invalidate_cache()


def _iter(monkeypatch, rows):
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    return list(_iter_rows_excel(_make_xlsx(rows)))


def _norm(monkeypatch, rows):
    from app.services.ingest import _normalize_row

    return [_normalize_row(r) for r in _iter(monkeypatch, rows)]


# ─────────────────────────────────────────────────────────────
# 1. _parse_simple_time 單元
# ─────────────────────────────────────────────────────────────
def test_parse_simple_time_minguo():
    from app.services.ingest import _parse_simple_time

    dt = _parse_simple_time("115/06/28 13:20")
    assert dt is not None
    assert (dt.year, dt.month, dt.day, dt.hour, dt.minute) == (2026, 6, 28, 13, 20)


def test_parse_simple_time_passthrough_formats():
    from app.services.ingest import _parse_simple_time

    assert _parse_simple_time("2026/06/28 13:20").year == 2026
    assert _parse_simple_time("2026-06-28 13:20:30").second == 30
    us = _parse_simple_time("6/28/2026 1:20 PM")          # 美式不可被誤當民國
    assert us is not None and (us.year, us.month, us.day, us.hour) == (2026, 6, 28, 13)


def test_parse_simple_time_bad():
    from app.services.ingest import _parse_simple_time

    assert _parse_simple_time("不是時間") is None
    assert _parse_simple_time("") is None
    assert _parse_simple_time(None) is None


# ─────────────────────────────────────────────────────────────
# 2. _parse_latlng_text 單元
# ─────────────────────────────────────────────────────────────
def test_parse_latlng_variants():
    from app.services.ingest import _parse_latlng_text

    assert _parse_latlng_text("22.6273,120.3014") == (22.6273, 120.3014)
    assert _parse_latlng_text("22.6273 120.3014") == (22.6273, 120.3014)
    assert _parse_latlng_text("22.6273，120.3014") == (22.6273, 120.3014)   # 全形逗號
    assert _parse_latlng_text("120.3014,22.6273") == (22.6273, 120.3014)   # lng,lat 自動對調


def test_parse_latlng_rejects():
    from app.services.ingest import _parse_latlng_text

    assert _parse_latlng_text("高雄市前金區中正四路211號") is None   # 含中文 → 地址
    assert _parse_latlng_text("211號") is None
    assert _parse_latlng_text("46697010805014") is None             # 單一數字（cell_id）
    assert _parse_latlng_text("1,2,3") is None                      # 非恰 2 浮點
    assert _parse_latlng_text("999,999") is None                    # 超界


# ─────────────────────────────────────────────────────────────
# 3. 有表頭「時間 / 位置」
# ─────────────────────────────────────────────────────────────
def test_headed_time_location_address(monkeypatch):
    rows = [
        ["時間", "位置"],
        ["2026/06/28 13:20", "高雄市前金區中正四路211號"],
        ["2026/06/28 13:25", "高雄市鼓山區明華路100號"],
    ]
    out = _norm(monkeypatch, rows)
    assert len(out) == 2
    assert out[0]["cell_addr"] == "高雄市前金區中正四路211號"
    assert "lat" not in out[0] or out[0].get("lat") in (None,)  # 地址走 geocode，無直給座標


# ─────────────────────────────────────────────────────────────
# 4. 無表頭：第一列就是資料
# ─────────────────────────────────────────────────────────────
def test_headerless_first_row_is_data(monkeypatch):
    rows = [
        ["2026/06/28 13:20", "高雄市前金區中正四路211號"],
        ["2026/06/28 13:25", "高雄市鼓山區明華路100號"],
    ]
    out = _norm(monkeypatch, rows)
    assert len(out) == 2, f"無表頭應讀到 2 列，實際 {len(out)}"
    assert out[0]["cell_addr"].startswith("高雄市前金區")


# ─────────────────────────────────────────────────────────────
# 5. B 欄 lat,lng / lng,lat / 全形 / 空白
# ─────────────────────────────────────────────────────────────
def test_b_latlng(monkeypatch):
    rows = [
        ["時間", "位置"],
        ["2026/06/28 13:20", "22.6273,120.3014"],   # lat,lng
        ["2026/06/28 13:21", "120.3014,22.6273"],   # lng,lat → 對調
        ["2026/06/28 13:22", "22.6273，120.3014"],  # 全形逗號
        ["2026/06/28 13:23", "22.6273 120.3014"],   # 空白分隔
    ]
    out = _norm(monkeypatch, rows)
    assert len(out) == 4
    for o in out:
        assert abs(float(o["lat"]) - 22.6273) < 1e-6
        assert abs(float(o["lng"]) - 120.3014) < 1e-6


# ─────────────────────────────────────────────────────────────
# 6. 民國年 + 美式時間（走完整路徑）
# ─────────────────────────────────────────────────────────────
def test_minguo_and_us_time_via_pipeline(monkeypatch):
    rows = [
        ["115/06/28 13:20", "22.6273,120.3014"],    # 民國年
        ["6/28/2026 1:20 PM", "22.7,120.4"],        # 美式
    ]
    from app.services.ingest import _iter_rows_excel, _parse_ts
    out = list(_iter_rows_excel(_make_xlsx(rows))) if False else _iter(monkeypatch, rows)
    assert len(out) == 2
    t0 = _parse_ts(out[0]["開始時間"])
    t1 = _parse_ts(out[1]["開始時間"])
    assert (t0.year, t0.month, t0.day, t0.hour, t0.minute) == (2026, 6, 28, 13, 20)
    assert (t1.year, t1.month, t1.day, t1.hour) == (2026, 6, 28, 13)


# ─────────────────────────────────────────────────────────────
# 7. A/B 以外雜欄被忽略
# ─────────────────────────────────────────────────────────────
def test_extra_columns_ignored(monkeypatch):
    rows = [
        ["時間", "位置", "備註", "序號"],
        ["2026/06/28 13:20", "高雄市前金區中正四路211號", "雜訊A", "001"],
        ["2026/06/28 13:25", "22.6273,120.3014", "雜訊B", "002"],
    ]
    out = _norm(monkeypatch, rows)
    assert len(out) == 2
    assert out[0]["cell_addr"] == "高雄市前金區中正四路211號"
    assert abs(float(out[1]["lat"]) - 22.6273) < 1e-6
    # 雜欄不應汙染 normalize 結果（無 備註/序號 對應的 canonical 欄）
    assert "備註" not in out[0] and "序號" not in out[0]


# ─────────────────────────────────────────────────────────────
# 8. 時間壞列 skipped，不讓整檔失敗
#    用「10 好列 + 1 壞時間列」證明：分母夠大時 1 壞列（10/11≈91%）不跌破 80% 門檻，
#    格式仍命中、壞列在 emit 階段被跳過 —— 不靠放寬門檻。
# ─────────────────────────────────────────────────────────────
def test_bad_time_row_skipped(monkeypatch):
    rows = [["時間", "位置"]]
    for i in range(10):
        rows.append([f"2026/06/28 13:{i:02d}", f"高雄市前金區中正四路{i+1}號"])
    rows.append(["不是時間", "高雄市鼓山區明華路100號"])   # 第 11 列壞時間 → emit 跳過
    out = _norm(monkeypatch, rows)
    assert len(out) == 10, f"10 好列 + 1 壞時間列 → 壞列跳過剩 10，實際 {len(out)}"


# ─────────────────────────────────────────────────────────────
# 9. 位置壞列不讓整檔失敗（B 空 → 仍 yield，下游視為未定位）
#    用「10 好列 + 1 空位置列」：位置命中 10/11≈91% ≥80% 仍命中；空位置列照 yield。
# ─────────────────────────────────────────────────────────────
def test_bad_location_not_fatal(monkeypatch):
    rows = [["時間", "位置"]]
    for i in range(10):
        rows.append([f"2026/06/28 14:{i:02d}", f"高雄市三民區建工路{i+1}號"])
    rows.append(["2026/06/28 14:59", ""])   # 第 11 列位置空 → 不致命
    out = _iter(monkeypatch, rows)
    assert len(out) == 11   # 11 列皆 yield（時間皆有效）；空位置列由下游判為未定位


# ─────────────────────────────────────────────────────────────
# 10. 不誤判電信格式（台哥大 / 雙向通聯 / 一般業者皆過規則 B、不進 fallback）
# ─────────────────────────────────────────────────────────────
def test_not_misdetect_tw_mobile(monkeypatch):
    # 台哥大上網歷程表頭（P8 別名）→ 過規則 B → 走正常 header 路徑
    rows = [
        ["進入基地台時間", "基地台停留時間", "離開基地台時間", "離開基地台編號", "離開基地台地址"],
        ["2025-10-01T00:03:22", "580", "2025-10-01T00:13:02", "466970108050142", "台北市萬華區長沙街二段188號"],
        ["2025-10-01T00:13:02", "108", "2025-10-01T00:14:50", "466970108050142", "台北市萬華區長沙街二段188號"],
    ]
    out = _iter(monkeypatch, rows)
    assert len(out) == 2
    assert "進入基地台時間" in out[0]      # 證明走正常路徑，非 simple（simple 會用「開始時間」）
    assert "開始時間" not in out[0]


def test_not_misdetect_general_format(monkeypatch):
    # 一般業者：開始時間 + 基地台編號 + 基地台地址（3 canonical 命中）→ 正常路徑
    rows = [
        ["開始時間", "基地台編號", "基地台地址"],
        ["2026/06/28 13:20", "CELL_A", "高雄市前金區中正四路211號"],
        ["2026/06/28 13:25", "CELL_B", "高雄市鼓山區明華路100號"],
    ]
    out = _iter(monkeypatch, rows)
    assert len(out) == 2
    assert "基地台編號" in out[0]          # 正常路徑保留 cell_id 欄
    assert "開始時間" in out[0]            # 此為原始表頭名，非 simple 注入


def test_simple_detection_not_triggered_on_telecom(monkeypatch):
    # 直接驗 _iter_simple_time_location 對電信 df_raw（B 欄非地址/座標）不命中
    import pandas as pd
    from app.services.ingest import _iter_simple_time_location

    df = pd.DataFrame([
        ["進入基地台時間", "基地台停留時間"],
        ["2025-10-01T00:03:22", "580"],
        ["2025-10-01T00:13:02", "108"],
    ])
    assert list(_iter_simple_time_location(df)) == []   # B 欄是秒數，非位置 → 不命中
