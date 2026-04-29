# backend/app/tests/test_ingest_buried_header.py
"""
W2.2 表頭埋深偵測單元測試（2026-04-29）

驗證 _iter_rows_excel 在「真表頭不在 row 0/1」情境下的新行為：
  1. 表頭在 row 0（純資料表）→ 向後相容
  2. 表頭在 row 1（W2.1 supertitle）→ 向後相容
  3. 表頭在 row 5 → W2.2 新能力
  4. 表頭在 row 22 → W2.2 極限案例（嫌1 網路歷程）
  5. row 10/11 重複表頭 → 取首次（row 10），row 11 變偽資料由 ingest 端過濾
  6. scan 視窗內無夠強 header → 規則 B 跳過
  7. PII 在前、真表頭在後 → 取真表頭（命中分數高）

不依賴 DB；用 monkeypatch 強制走 _RAW2CANON fallback。
"""
from __future__ import annotations

import io
import os
import logging
from typing import Any, List, Tuple

# 必須在 import app.* 之前設好環境變數，避免 db.session 在 import 期間崩潰
os.environ.setdefault("DATABASE_URL", "postgresql://user:pass@localhost:5432/fakedb")
os.environ.setdefault("SECRET_KEY", "test-secret-key-only-for-pytest")
os.environ.setdefault("AUTH_ENABLED", "true")


# ─────────────────────────────────────────────────────────────
# Fixture：用 openpyxl 動態組 in-memory xlsx，header 可放任意 row
# ─────────────────────────────────────────────────────────────
def _make_xlsx_buried(
    sheets: List[Tuple[str, int, List[Any], List[List[Any]], List[List[Any]]]]
) -> bytes:
    """
    sheets = [(sheet_name, header_row_idx, header, pre_rows, data_rows), ...]
      - header_row_idx：0-based，表頭實際應出現在 Excel 的第幾列（0=第1列）
      - pre_rows：表頭之前的雜訊（metadata / PII / supertitle 等）
                  長度應 = header_row_idx
      - data_rows：表頭之後的資料

    驗證 _iter_rows_excel 能正確找到 header_row_idx，並只 yield data_rows。
    """
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, idx, hdr, pre, data in sheets:
        assert len(pre) == idx, f"pre_rows 長度({len(pre)}) 應 == header_row_idx({idx})"
        ws = wb.create_sheet(title=name)
        for r in pre:
            ws.append(r)
        ws.append(hdr)
        for r in data:
            ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _patch_active_map_to_default(monkeypatch):
    import app.services.carrier_profile as cp
    monkeypatch.setattr(cp, "_load_default_profile_from_db", lambda: None)
    cp.invalidate_cache()


# ─────────────────────────────────────────────────────────────
# 1. 表頭在 row 0：向後相容
# ─────────────────────────────────────────────────────────────
def test_header_at_row_0(monkeypatch):
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    blob = _make_xlsx_buried([
        ("通聯", 0,
         ["時間", "基地台", "通話對象"],
         [],  # 表頭就在 row 0
         [[f"2026-01-01 1{i}:00:00", f"CELL_{i}", "0911"] for i in range(5)]),
    ])
    rows = list(_iter_rows_excel(blob))
    assert len(rows) == 5
    assert rows[0]["時間"] == "2026-01-01 10:00:00"
    assert rows[0]["基地台"] == "CELL_0"


# ─────────────────────────────────────────────────────────────
# 2. 表頭在 row 1（supertitle / 行動上網大標）：向後相容
# ─────────────────────────────────────────────────────────────
def test_header_at_row_1_supertitle(monkeypatch):
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    blob = _make_xlsx_buried([
        ("網路", 1,
         ["啟始時間", "結束時間", "基地台ID", "基地台位址"],
         [["行動上網", None, None, None]],  # row 0 跨欄大標
         [[f"2026-01-01 1{i}:00:00", f"2026-01-01 1{i}:30:00",
           f"CID_{i}", f"台北市中山區1號-{i}"] for i in range(5)]),
    ])
    rows = list(_iter_rows_excel(blob))
    assert len(rows) == 5, "supertitle 模式應仍能正確讀到 5 列"
    assert "啟始時間" in rows[0]
    assert "_unnamed_" not in str(list(rows[0].keys())), (
        "不該有 Unnamed 欄殘留（演算法應正確識別 row 1 為 header）"
    )


# ─────────────────────────────────────────────────────────────
# 3. 表頭在 row 5：W2.2 核心新能力
# ─────────────────────────────────────────────────────────────
def test_header_at_row_5(monkeypatch):
    """模擬電話通聯+歷程「嫌2 雙向歷程」：row 0-3 查詢條件、row 4 真表頭"""
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    pre = [
        ["查詢條件", None, None, None, None],
        ["通聯類別: 雙向", "查詢狀態: 成功", None, None, None],
        ["區段時間: 2024-09-01", None, None, None, None],
        ["電話號碼:0972124190", None, None, None, None],
    ]
    data = [[f"2024-09-04 08:4{i}:00", "1", "972124190", "722014032",
             "高雄市左營區新庄仔路515巷21號"] for i in range(8)]
    blob = _make_xlsx_buried([
        ("嫌2雙向", 4,
         ["始話時間", "通話秒數", "調閱號碼", "基地台/交換機", "備註"],
         pre, data),
    ])
    rows = list(_iter_rows_excel(blob))
    assert len(rows) == 8, f"應只讀資料 8 列，metadata 不該進入；實際 {len(rows)}"
    assert rows[0]["始話時間"] == "2024-09-04 08:40:00"
    assert rows[0]["基地台/交換機"] == "722014032"
    # metadata 欄不該殘留
    assert "查詢條件" not in rows[0]
    assert "通聯類別: 雙向" not in rows[0]


# ─────────────────────────────────────────────────────────────
# 4. 表頭在 row 22：W2.2 極限案例
# ─────────────────────────────────────────────────────────────
def test_header_at_row_22(monkeypatch):
    """
    模擬電話通聯+歷程「嫌1 網路歷程」極端結構：
    row 0-20 是查詢條件 + 用戶資訊 PII，row 21 是「行動上網」supertitle，
    row 22 才是真表頭。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    # 21 列雜訊 + 1 列 supertitle
    pre = []
    for i in range(21):
        pre.append([f"meta_{i}", None, None, None])
    pre.append(["行動上網", None, None, None])

    data = [[f"2024-08-01T0{i}:50:30", str(7200 + i),
             f"CID_{i}", f"高雄市左營區至聖路200號-{i}"] for i in range(6)]
    blob = _make_xlsx_buried([
        ("嫌1網路", 22,
         ["啟始時間", "通聯時間(秒)", "基地台ID", "基地台位址"],
         pre, data),
    ])
    rows = list(_iter_rows_excel(blob))
    assert len(rows) == 6, f"應讀資料 6 列；實際 {len(rows)}"
    assert "啟始時間" in rows[0]
    assert rows[0]["啟始時間"] == "2024-08-01T00:50:30"


# ─────────────────────────────────────────────────────────────
# 5. row 10/11 重複表頭：取首次出現
# ─────────────────────────────────────────────────────────────
def test_duplicated_header_takes_first(monkeypatch):
    """
    模擬嫌1 雙向歷程：row 10 = 表頭、row 11 = 表頭（重複）、row 12+ 是資料。
    演算法應取 row 10；row 11 變成偽資料列（內容是表頭文字）。
    這個偽列的 start_ts 會是字串「始話時間」，被 ingest 端 _parse_ts 過濾。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel, _normalize_row, _parse_ts

    # 9 列雜訊 → row 10（idx=9）真表頭 → row 11（idx=10）重複表頭文字（變偽資料）→ row 12+ 真資料
    pre = [[f"meta_{i}", None, None] for i in range(9)]
    data = [
        ["始話時間", "通話秒數", "基地台/交換機"],  # 第 1 列 data：偽表頭
    ]
    data += [[f"2024-09-04 08:4{i}:00", "1", "722014032"] for i in range(5)]

    blob = _make_xlsx_buried([
        ("嫌1雙向", 9,
         ["始話時間", "通話秒數", "基地台/交換機"],
         pre, data),  # pre 正好 9 列、header 第 10 列、data 6 列（1 偽 + 5 真）
    ])

    rows = list(_iter_rows_excel(blob))
    assert len(rows) == 6, f"應讀 6 列（1 偽 + 5 真）；實際 {len(rows)}"

    # 跑 normalize + parse_ts，看真實能 pass 幾列
    ok = 0
    for raw in rows:
        norm = _normalize_row(raw)
        if _parse_ts(norm.get("start_ts")):
            ok += 1
    assert ok == 5, f"5 列真資料能 parse；偽 header 列被自動過濾。實際 ok={ok}"


# ─────────────────────────────────────────────────────────────
# 6. scan 視窗內無夠強 header → 規則 B 跳過
# ─────────────────────────────────────────────────────────────
def test_no_real_header_in_window(monkeypatch, caplog):
    """
    所有 25 列都是雜訊或 PII，找不到 >= 2 命中的 header。
    應跳過該 sheet 並寫入 log。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    # 30 列全是 PII（姓名/出生/電話/身分證 — 一個都不命中 cell_id/cell_addr/start_ts）
    junk = [["姓名", "出生", "電話", "身分證"]]
    junk += [[f"王小{i}", "1990", "0911", f"A{i:09d}"] for i in range(29)]

    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="人資")
    for r in junk:
        ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    blob = bio.getvalue()

    with caplog.at_level(logging.INFO, logger="app.services.ingest"):
        rows = list(_iter_rows_excel(blob))
    assert rows == []
    assert any("人資" in m and "header matches" in m for m in caplog.messages), (
        f"未在 log 看到「人資」sheet 被以 header matches 不足為由跳過；"
        f"實際 log: {caplog.messages}"
    )


# ─────────────────────────────────────────────────────────────
# 7. PII 早期命中 1 欄、真表頭命中 6 欄 → 演算法取真表頭
# ─────────────────────────────────────────────────────────────
def test_pii_header_loses_to_real_header(monkeypatch):
    """
    模擬嫌1 雙向歷程的核心挑戰：
      row 5: ['個人資料', '電話號碼', '地址', '聯絡電話', '使用期間', '備註']
              → 「地址」會命中 cell_addr（1 個命中）
      row 10: ['始話時間', '通話秒數', '調閱號碼', 'IMEI', '通話類別',
               '通話對象', '轉接電話', '基地台/交換機']
              → 命中 始話時間→start_ts、基地台/交換機→cell_id（2 個命中）
    演算法應取 row 10（命中分數較高），不是 row 5。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    pre = [
        ["查詢條件", None, None, None, None, None, None, None],
        ["通聯類別: 雙向", None, None, None, None, None, None, None],
        ["區段時間", None, None, None, None, None, None, None],
        ["電話號碼", None, None, None, None, None, None, None],
        # row 5（idx=4）：偽表頭「個人資料/電話號碼/地址/...」
        ["個人資料", "電話號碼", "地址", None, "聯絡電話", "使用期間", "備註", None],
        # row 6-9（idx=5-8）：個人資料 PII
        ["身份證號碼: A123", "0911111111", "戶籍: 高雄市", None, "0922222222", "2023-10", "預付", None],
        ["姓名: 洪某", None, "帳寄: 高雄市", None, None, "正常", None, None],
        ["性別: 女", None, None, None, None, None, None, None],
        ["出生: 1993", None, None, None, None, None, None, None],
    ]
    data = [[f"2024-08-0{i+1}T03:54:00", "1", "85293633906", "356087099882710",
             "收簡訊", "2353494D", "0", "717511031"] for i in range(6)]

    blob = _make_xlsx_buried([
        ("嫌1雙向", 9,  # 真表頭在 row 10（idx=9）
         ["始話時間", "通話秒數", "調閱號碼", "IMEI", "通話類別", "通話對象", "轉接電話", "基地台/交換機"],
         pre, data),
    ])
    rows = list(_iter_rows_excel(blob))
    assert len(rows) == 6, f"應只讀真資料 6 列；實際 {len(rows)}"
    # 真表頭欄位應出現
    assert "始話時間" in rows[0]
    assert "基地台/交換機" in rows[0]
    # 偽表頭欄位（個人資料/電話號碼）不該成為 dict key
    assert "個人資料" not in rows[0]
    assert "電話號碼" not in rows[0]
    # PII metadata 不該進入（被切掉在 header 之上）
    for r in rows:
        for v in r.values():
            assert "身份證號碼" not in str(v)
            assert "姓名: " not in str(v)


# ─────────────────────────────────────────────────────────────
# 8. ingest 整合：W2.2 表頭找對 + W1.5 空值 fallback + 方言別名
# ─────────────────────────────────────────────────────────────
def test_w2_2_dialect_with_normalize(monkeypatch):
    """
    模擬電話通聯+歷程「嫌2 網路歷程」schema：
      row 0-3 metadata
      row 4 真表頭：['資料來源', '檔案序號', '交換機代號', '用戶號碼',
                    'IMEI', '手機連到基地台的時間', '秒數',
                    '連到internet的時間', '...', 'IP', '基地台代碼', '基地台地址']
      row 5+ 資料：「手機連到基地台的時間」全空（此 carrier 怪癖），
                  真實時間在「連到internet的時間」欄

    驗證：W2.2 找到 row 4 表頭 + W2.2 新增別名（連到internet的時間 →
    start_ts、基地台代碼 → cell_id）+ W1.5 空值 fallback 共同生效。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel, _normalize_row, _parse_ts

    pre = [
        [None, "查詢條件", None, None, None, None, None, None, None, None, None, None],
        [None, "通聯類別: 發信", None, None, None, None, None, None, None, None, None, None],
        [None, "區段時間: 2024-09-01", None, None, None, None, None, None, None, None, None, None],
        [None, "電話號碼:886972124190", None, None, None, None, None, None, None, None, None, None],
    ]
    data = [
        ["85", "ipv4", "3L1EPG3_PGW", "972124190", "352976240414080",
         "",  # 手機連到基地台的時間 → 空（此 carrier 怪癖）
         "0",
         f"2024-09-04 11:24:4{i}",  # 連到internet的時間 → 真實時間
         "3647", "10.142.253.227",
         f"31326201{i}",  # 基地台代碼
         "桃園市中壢區康樂路77號"]
        for i in range(5)
    ]
    blob = _make_xlsx_buried([
        ("嫌2網路", 4,
         ["資料來源", "檔案序號", "交換機代號", "用戶號碼", "用戶手機序號IMEI",
          "手機連到基地台的時間", "手機連到基地台的秒數",
          "連到internet的時間", "連到internet的秒數",
          "用戶連線時被指配之內網IP", "基地台代碼", "基地台地址"],
         pre, data),
    ])
    rows = list(_iter_rows_excel(blob))
    assert len(rows) == 5, f"應讀 5 列；實際 {len(rows)}"

    ok = 0
    for raw in rows:
        norm = _normalize_row(raw)
        st = _parse_ts(norm.get("start_ts"))
        cid = (str(norm.get("cell_id") or "").strip() or None)
        addr = (str(norm.get("cell_addr") or "").strip() or None)
        # W1.5 空值 fallback：手機連到基地台的時間（空）不該蓋掉
        # 連到internet的時間（有值）
        if st and cid and addr:
            ok += 1
    assert ok == 5, (
        f"全部 5 列都應正確 normalize（W2.2 表頭 + 別名 + W1.5 fallback 三方協作）；"
        f"實際 ok={ok}"
    )
