# backend/app/tests/test_ingest_tw_mobile_data.py
"""
台哥大「通訊數據上網歷程」格式支援 + 假 dimension 防呆（2026-06-27，test2.xlsx）

Background
==========
使用者上傳台哥大上網歷程 `test2.xlsx`（21,785 列）回 422「讀不到」。完整走
`_iter_rows_excel` 後查出**兩個獨立關卡**疊加：

  1. 真表頭埋在 row 27（前面是查詢條件 + 完整「使用者資料」PII 區塊：用戶名稱 /
     帳寄地址 / 戶籍地址 / 證號 / 生日 …）。舊 SCAN_WINDOW=25 掃不到 → 規則 B
     judge best_match=0 → 整張 sheet 被當非資料表跳過 → yield 0 列。
  2. 此 carrier 的欄名是「進入/離開基地台」系列，全不在 _RAW2CANON：
       進入基地台時間 / 離開基地台時間 / 離開基地台編號 / 離開基地台地址
     即使掃到表頭也 0 命中。

附帶查到 openpyxl 假 dimension bug：此檔 worksheet 的 <dimension> 被匯出工具寫死
成 'A1'，openpyxl read_only 模式信任它 → _peek_headers/_peek_sample_rows（診斷／
手動對應 UI 用）只讀到 1 列、範例值全空 → 手動對應介面無法運作。

本檔三組測試各鎖一個關卡，避免回歸。不依賴 DB（monkeypatch carrier_profile）。
"""
from __future__ import annotations

import io
import os
import re
import zipfile
from typing import Any, List

os.environ.setdefault("DATABASE_URL", "postgresql://user:pass@localhost:5432/fakedb")
os.environ.setdefault("SECRET_KEY", "test-secret-key-only-for-pytest")
os.environ.setdefault("AUTH_ENABLED", "true")


# ─────────────────────────────────────────────────────────────
# 共用：組 in-memory xlsx；可選擇把 <dimension> 竄改成 'A1'
# ─────────────────────────────────────────────────────────────
def _make_xlsx(rows: List[List[Any]], fake_dimension_a1: bool = False) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工作表1"
    for r in rows:
        ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    data = bio.getvalue()
    if not fake_dimension_a1:
        return data

    # 把 sheet1.xml 的 <dimension ref="..."/> 竄改成 'A1'，重現台哥大匯出工具
    # 寫死假邊界的情境（openpyxl read_only 會信任它而只讀到 1 列）。
    src = zipfile.ZipFile(io.BytesIO(data))
    out_bio = io.BytesIO()
    with zipfile.ZipFile(out_bio, "w", zipfile.ZIP_DEFLATED) as dst:
        for name in src.namelist():
            blob = src.read(name)
            if name == "xl/worksheets/sheet1.xml":
                txt = blob.decode("utf-8")
                txt = re.sub(r'<dimension ref="[^"]*"/>', '<dimension ref="A1"/>', txt)
                blob = txt.encode("utf-8")
            dst.writestr(name, blob)
    return out_bio.getvalue()


# 台哥大上網歷程真實表頭（10 欄）
_TW_HEADER = [
    "進入基地台時間", "基地台停留時間", "離開基地台時間",
    "離開基地台編號", "離開基地台地址",
    "上傳使用量(Byte)", "下載使用量(Byte)", "全部使用量(Byte)", "IMEI", "備註",
]


def _tw_data_row(i: int) -> List[Any]:
    return [
        f"2025-10-01T0{i}:03:22", "580", f"2025-10-01T0{i}:13:02",
        "466970108050142", "台北市萬華區長沙街二段188號6樓頂",
        "859523", "6147949", "7007472", "357507335416570", "備註",
    ]


def _patch_active_map_to_default(monkeypatch):
    """讓 _iter_rows_excel 在規則 B 用 _RAW2CANON，避免測試碰 DB。"""
    import app.services.carrier_profile as cp

    monkeypatch.setattr(cp, "_load_default_profile_from_db", lambda: None)
    cp.invalidate_cache()


# ─────────────────────────────────────────────────────────────
# 1. 欄名別名：_normalize_row 正確映射「進入/離開基地台」系列
# ─────────────────────────────────────────────────────────────
def test_tw_mobile_aliases_normalize():
    from app.services.ingest import _normalize_row

    raw = {
        "進入基地台時間": "2025-10-01T00:03:22",
        "離開基地台時間": "2025-10-01T00:13:02",
        "離開基地台編號": "466970108050142",
        "離開基地台地址": "台北市萬華區長沙街二段188號6樓頂",
        "基地台停留時間": "580",
    }
    n = _normalize_row(raw)
    assert n["start_ts"] == "2025-10-01T00:03:22"   # 進入 → 定位時間
    assert n["end_ts"] == "2025-10-01T00:13:02"     # 離開 → end_ts（不遺失）
    assert n["cell_id"] == "466970108050142"        # 純 ID，非複合欄
    assert n["cell_addr"] == "台北市萬華區長沙街二段188號6樓頂"


# ─────────────────────────────────────────────────────────────
# 2. 深埋表頭：真表頭在 row 27（超過舊 SCAN_WINDOW=25）仍須讀到
# ─────────────────────────────────────────────────────────────
def test_tw_mobile_deep_buried_header(monkeypatch):
    """
    重現 test2.xlsx：27 列前置（查詢條件 + PII 使用者資料）後才是真表頭。
    舊 SCAN_WINDOW=25 會 yield 0；放寬到 30 後應讀到全部資料列且映射正確。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel, _normalize_row

    preamble = [
        ["台灣大哥大通訊數據上網歷程查詢"],
        ["回覆時間:2026-06-24T09:00:00"], ["回覆狀態:成功"], [],
        ["查詢條件"], [],
        ["(1)調閱目標類別:MSISDN"], ["業務種類:3G"], ["使用者資料"],
        ["用戶識別碼:"], ["用戶名稱:余紹丞"], ["帳寄地址:新北市蘆洲區集賢路"],
        ["戶籍地址:247新北市蘆洲區"], ["證號類別1:未知"], ["證號類別2:未知"],
        ["性別:男性"], ["生日:2000-06-06"], ["申請日期:2022-04-15"],
        ["生效起始時間:2022-04-15"], ["備註:服務狀態：正常使用中。"],
        ["申請號碼:0908008051"], ["門號類型:一般卡"], ["其他聯絡電話:0282863"],
        ["聯絡電話:0282863741"], ["其他聯絡電話用途:"], [],
        ["通聯之基地台相關資訊"],
    ]
    assert len(preamble) == 27   # 真表頭將落在 index 27
    rows_in = preamble + [_TW_HEADER] + [_tw_data_row(i) for i in range(1, 6)]

    blob = _make_xlsx(rows_in)
    out = list(_iter_rows_excel(blob))
    assert len(out) == 5, f"應讀到 5 筆資料，實際 {len(out)}"

    n0 = _normalize_row(out[0])
    assert n0["cell_id"] == "466970108050142"
    assert n0["cell_addr"].startswith("台北市萬華區")
    assert n0["start_ts"].startswith("2025-10-01T01")


def test_scan_window_just_reaches_header(monkeypatch):
    """守住放寬幅度：表頭在 index 29（SCAN_WINDOW=30 的邊界）仍須讀到。"""
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    preamble = [[f"metadata:{i}"] for i in range(29)]   # index 0..28
    rows_in = preamble + [_TW_HEADER] + [_tw_data_row(1), _tw_data_row(2)]
    out = list(_iter_rows_excel(_make_xlsx(rows_in)))
    assert len(out) == 2


# ─────────────────────────────────────────────────────────────
# 3. 假 dimension 防呆：peek 路徑（手動對應 UI 用）仍讀得到範例值
# ─────────────────────────────────────────────────────────────
def test_read_xlsx_top_rows_fallback_triggers(monkeypatch):
    """
    確定性驗證 fallback 分支：當 read_only 因假 dimension 回「多列但每列幾乎全空」
    （最寬列 ≤1 格 = 退化徵狀），_read_xlsx_top_rows 必須改用 read_only=False 重讀。

    用 monkeypatch 模擬此退化（real test2.xlsx 實證：read_only 35 列最寬僅 1 格），
    避免依賴特定匯出工具才會觸發的 openpyxl 行為（合成檔不一定重現）。
    """
    import app.services.ingest as ing
    import openpyxl

    rows_in = [_TW_HEADER] + [_tw_data_row(i) for i in range(1, 5)]
    blob = _make_xlsx(rows_in)  # 正常 dimension；靠 monkeypatch 模擬退化

    calls = {"read_only_true": 0, "read_only_false": 0}
    orig = openpyxl.load_workbook

    def fake_load(buf, *, read_only, data_only):
        if read_only:
            calls["read_only_true"] += 1
            # 模擬假 dimension：read_only 回多列，但每列只剩第 1 格（其餘讀成空）
            wb = orig(buf, read_only=False, data_only=data_only)

            class _DegenerateWS:
                def __init__(self, ws):
                    self._ws = ws

                def iter_rows(self, **kw):
                    for row in self._ws.iter_rows(**kw):
                        yield (row[0],) + (None,) * (len(row) - 1)

            class _Wb:
                active = _DegenerateWS(wb.active)

                def close(self):
                    wb.close()

            return _Wb()
        calls["read_only_false"] += 1
        return orig(buf, read_only=False, data_only=data_only)

    monkeypatch.setattr(openpyxl, "load_workbook", fake_load)

    top = ing._read_xlsx_top_rows(blob, 4)
    assert calls["read_only_true"] == 1, "應先試 read_only 快路徑"
    assert calls["read_only_false"] == 1, "退化（最寬列 ≤1）應觸發 read_only=False fallback"
    assert len(top) >= 2, f"fallback 應讀到多列，實際 {len(top)}"
    assert top[0][0] == "進入基地台時間"


def test_peek_headers_and_sample_rows_on_fake_dimension():
    """手動對應 UI 依賴的 _peek_headers / _peek_sample_rows 在假 dimension 檔仍可用。"""
    from app.services.ingest import _peek_headers, _peek_sample_rows

    rows_in = [_TW_HEADER] + [_tw_data_row(i) for i in range(1, 5)]
    blob = _make_xlsx(rows_in, fake_dimension_a1=True)

    headers = _peek_headers("test.xlsx", blob)
    assert headers[:5] == _TW_HEADER[:5], f"表頭抓取失敗：{headers}"

    samples = _peek_sample_rows("test.xlsx", blob, n=3)
    assert len(samples) == 3, f"應抓到 3 列範例值，實際 {len(samples)}"
    # 範例值裡應看得到時間與地址（手動對應靠這些猜欄）
    assert any("2025-10-01" in c for c in samples[0])
    assert any("台北市萬華區" in c for c in samples[0])
