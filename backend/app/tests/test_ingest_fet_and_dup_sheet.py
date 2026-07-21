# backend/app/tests/test_ingest_fet_and_dup_sheet.py
"""
遠傳上網歷程格式 + 更深表頭 + 重複分頁去重（2026-07-21）

Background
==========
對 `基地台位置範例檔案/` 全 17 檔實測，3 個真實檔解析失敗（另 1 個是刻意的壞檔）。
根因是**三個互相獨立**的關卡，各自對應本檔一組測試：

  1. 遠傳「028351 / 031543.xlsx」欄名 `通聯起始時間` / `通聯結束時間` 不在
     _RAW2CANON。表頭在 row 23（窗內、有讀到 9319 列），但每列都拿不到
     start_ts → 全數被時間驗證濾掉 → 0 筆 → 422。
     另注意此 carrier 的資料落點：`起始基地台編號/地址` 兩欄**實測為空**、
     真值在 `離開基地台編號/地址`（後者 P8 已有別名）。兩組別名並存時，靠
     _normalize_row 的 W1.5「空值不覆蓋」語意互補，順序無關。

  2. 台哥大「026962 陳2號機網路.xlsx」真表頭埋在 **row 48**：該檔含「兩個調閱
     區塊」，每塊各帶一整段「使用者資料」PII，把表頭推得比 P8 的 test2.xlsx
     （row 27）更深。舊 SCAN_WINDOW=30 在窗內找不到任何命中列 → 規則 B 誤判
     整張 sheet 為非資料表 → yield 0 列。

  3. 「複本 029935」「031543」各含一張名為「標記」、與「工作表1」**逐格完全
     相同**的分頁（承辦人複製一份用來標記重點）。W2.1 多 sheet 支援會兩張都讀
     → 同一筆通聯入庫兩次；raw_traces 沒有內容層級唯一索引，DB 不會擋 → 地圖
     點位加倍、證物報告筆數加倍。這是證據完整性問題：「該時段出現幾次」是實質
     待證事實，翻倍會扭曲軌跡密度與停留時間的判讀。

依 W2.3-O 的踩雷教訓，2/3 兩組**走完整 `_iter_rows_excel` 路徑**（不只測
_normalize_row 單列），因為當初的 bug 正是卡在 sheet 層守門而非 normalize。
不依賴 DB（monkeypatch carrier_profile）。
"""
from __future__ import annotations

import io
import os
from typing import Any, Dict, List

os.environ.setdefault("DATABASE_URL", "postgresql://user:pass@localhost:5432/fakedb")
os.environ.setdefault("SECRET_KEY", "test-secret-key-only-for-pytest")
os.environ.setdefault("AUTH_ENABLED", "true")


# ─────────────────────────────────────────────────────────────
# 共用工具
# ─────────────────────────────────────────────────────────────
def _make_xlsx(sheets: Dict[str, List[List[Any]]]) -> bytes:
    """組 in-memory xlsx；dict 保序 → sheet 順序即宣告順序。"""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for r in rows:
            ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _patch_active_map_to_default(monkeypatch):
    """讓 _iter_rows_excel 在規則 B 用 _RAW2CANON，避免測試碰 DB。"""
    import app.services.carrier_profile as cp

    monkeypatch.setattr(cp, "_load_default_profile_from_db", lambda: None)
    cp.invalidate_cache()


# 遠傳上網歷程真實表頭（16 欄，節錄前 7 欄後補足）
_FET_HEADER = [
    "通聯起始時間", "通聯時間(秒)", "通聯結束時間",
    "起始基地台編號", "離開基地台編號",
    "起始基地台地址", "離開基地台地址",
    "外部IP位址", "內部IP位址", "手機序號",
]


def _fet_data_row(i: int) -> List[Any]:
    """重現實測落點：起始端兩欄為空、真值在離開端。"""
    return [
        f"2026-04-01T0{i}:24:26", "2700", f"2026-04-01T0{i}:69:26".replace(":69:", ":54:"),
        "",                                   # 起始基地台編號（空）
        "46601198530720092012",               # 離開基地台編號（真值）
        "",                                   # 起始基地台地址（空）
        "高雄市三民區本館里昌裕街1號7樓樓頂(4G)",  # 離開基地台地址（真值）
        "27.247.141.155:61504-65535", "", "3576841358505305",
    ]


# ─────────────────────────────────────────────────────────────
# 1. 遠傳欄名別名：_normalize_row 正確映射
# ─────────────────────────────────────────────────────────────
def test_fet_aliases_normalize():
    from app.services.ingest import _normalize_row

    n = _normalize_row({
        "通聯起始時間": "2026-04-01T00:24:26",
        "通聯結束時間": "2026-04-01T01:09:26",
        "起始基地台編號": "46601198530720092012",
        "起始基地台地址": "高雄市三民區本館里昌裕街1號7樓樓頂(4G)",
        "通聯時間(秒)": "2700",
    })
    assert n["start_ts"] == "2026-04-01T00:24:26"   # 通聯起始 → 定位時間
    assert n["end_ts"] == "2026-04-01T01:09:26"     # 通聯結束 → end_ts（不遺失）
    assert n["cell_id"] == "46601198530720092012"
    assert n["cell_addr"] == "高雄市三民區本館里昌裕街1號7樓樓頂(4G)"


def test_fet_empty_start_side_does_not_clobber_end_side():
    """
    起始端為空、離開端有值時，兩組別名都映射到 cell_id/cell_addr。
    W1.5「空值不覆蓋」必須讓真值勝出，且**與 dict 走訪順序無關**
    （空值在前、在後各測一次）。
    """
    from app.services.ingest import _normalize_row

    base = {
        "通聯起始時間": "2026-04-01T00:24:26",
        "離開基地台編號": "46601198530720092012",
        "離開基地台地址": "高雄市三民區本館里昌裕街1號7樓樓頂(4G)",
    }
    # 空值在後
    n1 = _normalize_row({**base, "起始基地台編號": "", "起始基地台地址": "   "})
    # 空值在前
    n2 = _normalize_row({"起始基地台編號": "", "起始基地台地址": "   ", **base})
    for n in (n1, n2):
        assert n["cell_id"] == "46601198530720092012"
        assert n["cell_addr"] == "高雄市三民區本館里昌裕街1號7樓樓頂(4G)"


def test_fet_full_excel_path(monkeypatch):
    """走完整 _iter_rows_excel：表頭埋在 row 23，須讀到全部資料列且映射正確。"""
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel, _normalize_row

    preamble = [
        ["遠傳通訊數據上網歷程查詢"],
        ["回覆時間:2026-06-24T19:12:14", "發文案號:028351-FARIP"],
        ["回覆狀態:成功", "回覆備註:"], [], [], [],
        ["(1)調閱目標類別:MSISDN(門號ISDN號碼)", "調閱目標:"],
        ["查詢起始時間:2026-04-0100:00:00"], ["查詢終止時間:2026-04-3023:59:59"],
        [], ["使用者資料"], ["啟用時間:2022-11-15T00:00:00"], ["用戶名稱:王莉思"],
        ["申請號碼:0903600077"], ["帳寄地址:高雄市三民區民祥街８號１１樓之２"],
        ["戶籍地址:高雄市三民區民祥街８號１１樓之２"], ["電子郵件:"],
        ["其他聯絡電話:0927597722"], ["證號類別1:身份證字號"], ["生日:1984-09-28"],
        ["申請日期:"], ["備註:"], ["基地台上網紀錄資料"],
    ]
    assert len(preamble) == 23  # 真表頭落在 index 23，與實檔一致

    rows = preamble + [_FET_HEADER] + [_fet_data_row(i) for i in range(1, 6)]
    got = list(_iter_rows_excel(_make_xlsx({"工作表1": rows})))
    assert len(got) == 5, f"應讀到 5 筆資料列，實得 {len(got)}"

    norm = [_normalize_row(r) for r in got]
    assert all(n.get("start_ts") for n in norm), "每列都必須有 start_ts（原 bug 全為 None）"
    assert all(n["cell_id"] == "46601198530720092012" for n in norm)
    assert all(n["cell_addr"].startswith("高雄市三民區") for n in norm)

    # forensic data minimization：表頭之上的 PII 不得外流
    blob = repr(norm)
    assert "王莉思" not in blob and "1984-09-28" not in blob


# ─────────────────────────────────────────────────────────────
# 2. 更深的埋藏表頭：兩個調閱區塊 → 真表頭在 row 48
# ─────────────────────────────────────────────────────────────
def _tw_block(idx: int) -> List[List[Any]]:
    """單一調閱區塊的前置段（查詢條件 + 一整段使用者資料 PII），共 21 列。"""
    return [
        [f"({idx})調閱目標類別:MSISDN(門號ISDN號碼)", "調閱目標:886908033327"],
        ["業務種類:3G", "業者名稱:台灣大哥大"],
        ["使用者資料"], ["用戶識別碼:"], ["用戶名稱:王世明"],
        ["帳寄地址:高雄市湖內區中山路二段226巷15弄5號"],
        ["戶籍地址:829高雄市湖內區中山路二段226巷15弄5號"],
        ["證號類別1:未知", "證號1:E120906008"],
        ["證號類別2:未知", "證號2:000091153330"],
        ["性別:男性"], ["生日:1969-01-12"], ["申請日期:2025-09-21"],
        ["生效起始時間:2025-09-21T00:00:00"], ["備註:服務狀態：正常使用中。"],
        ["申請號碼:0908033327"], ["門號類型:一般卡(台灣大哥大4G)"],
        ["其他聯絡電話:"], ["聯絡電話:"], ["其他聯絡電話用途:"], [], [],
    ]


_TW_HEADER = [
    "進入基地台時間", "基地台停留時間", "離開基地台時間",
    "離開基地台編號", "離開基地台地址",
]


def test_tw_mobile_header_row48_two_query_blocks(monkeypatch):
    """
    重現「026962 陳2號機網路.xlsx」：兩個調閱區塊把真表頭推到 row 48。
    舊 SCAN_WINDOW=30 → yield 0 列（規則 B 誤判為非資料表）；
    放寬到 60 後應讀到全部資料列。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel, _normalize_row

    rows: List[List[Any]] = [
        ["台灣大哥大通訊數據上網歷程查詢"],
        ["回覆時間:2026-06-25T14:08:09", "發文案號:026962"],
        ["回覆狀態:成功", "回覆備註:"], [], [], [],
    ]
    rows += _tw_block(1)
    rows += _tw_block(2)
    assert len(rows) == 48, f"真表頭應落在 index 48，實際前置 {len(rows)} 列"

    rows.append(_TW_HEADER)
    rows += [
        [f"2026-05-0{i}T23:37:44", "2700", f"2026-05-0{i}T23:59:44",
         "466970829017182", "高雄市湖內區中山路２段290號"]
        for i in range(1, 5)
    ]

    got = list(_iter_rows_excel(_make_xlsx({"工作表1": rows})))
    assert len(got) == 4, f"表頭埋在 row 48 仍須讀到 4 筆，實得 {len(got)}"

    norm = [_normalize_row(r) for r in got]
    assert all(n["cell_id"] == "466970829017182" for n in norm)
    assert all(n.get("start_ts") for n in norm)
    # 兩個區塊的 PII 都不得流出
    assert "王世明" not in repr(norm) and "E120906008" not in repr(norm)


def test_scan_window_covers_row48():
    """把 SCAN_WINDOW 的實際值鎖住 —— 它是 local 常數，只能從原始碼讀。

    為什麼要這條：先前驗證時曾用 monkeypatch 改 ingest.SCAN_WINDOW 而毫無作用
    （它是 _iter_rows_excel 的 local 變數，不是 module 屬性），差點據此誤判
    「放寬窗寬無效」。這條測試讓「窗寬 ≥ 49」這個前提被明確固定住。
    """
    import inspect
    import re

    from app.services.ingest import _iter_rows_excel

    src = inspect.getsource(_iter_rows_excel)
    m = re.search(r"^\s*SCAN_WINDOW\s*=\s*(\d+)", src, re.M)
    assert m, "找不到 SCAN_WINDOW 宣告"
    assert int(m.group(1)) >= 49, (
        f"SCAN_WINDOW={m.group(1)} 掃不到 026962 埋在 row 48 的真表頭"
    )


# ─────────────────────────────────────────────────────────────
# 3. 重複分頁去重（規則 A2）
# ─────────────────────────────────────────────────────────────
def _dup_sheet_rows() -> List[List[Any]]:
    return [_TW_HEADER] + [
        [f"2026-06-0{i}T10:00:00", "2700", f"2026-06-0{i}T10:45:00",
         "466970829017141", "高雄市湖內區中山路２段290號"]
        for i in range(1, 4)
    ]


def test_identical_sheet_is_deduplicated(monkeypatch):
    """
    「標記」與「工作表1」逐格相同 → 只採第一張，資料不得翻倍。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    rows = _dup_sheet_rows()
    single = list(_iter_rows_excel(_make_xlsx({"工作表1": rows})))
    both = list(_iter_rows_excel(_make_xlsx({"工作表1": rows, "標記": [r[:] for r in rows]})))

    assert len(single) == 3
    assert len(both) == 3, f"重複分頁應被去重，實得 {len(both)} 筆（翻倍即回歸）"


def test_near_duplicate_sheet_is_kept(monkeypatch):
    """
    只要有一格不同就當兩份獨立證據保留 —— 寧可多存不可漏存
    （漏存是不可逆的證據滅失，多存至少可事後篩選）。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    rows = _dup_sheet_rows()
    variant = [r[:] for r in rows]
    variant[1][4] = "高雄市三民區陽明路170巷8號11樓之一"  # 改一格地址

    got = list(_iter_rows_excel(_make_xlsx({"工作表1": rows, "另一次調閱": variant})))
    assert len(got) == 6, f"內容不同的分頁必須都保留，實得 {len(got)} 筆"


def test_dedup_does_not_merge_across_files(monkeypatch):
    """
    去重範圍限於「單一 workbook 內」—— 不同檔案含相同分頁必須各自完整解析
    （同一批資料由不同單位交付時，兩份都是獨立證物）。
    """
    _patch_active_map_to_default(monkeypatch)
    from app.services.ingest import _iter_rows_excel

    rows = _dup_sheet_rows()
    blob = _make_xlsx({"工作表1": rows})
    assert len(list(_iter_rows_excel(blob))) == 3
    assert len(list(_iter_rows_excel(blob))) == 3, "跨檔呼叫不得共用去重狀態"
