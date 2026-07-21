# backend/app/tests/test_ingest_subset_sheet_dedup.py
"""
規則 A2：同一 workbook 內的「multiset 子集分頁」去重（2026-07-21）

Background
==========
承辦人常把資料分頁複製一份用來標記重點。實測「031543 蘇網路.xlsx」的第二張分頁
是第一張的**純子集**（4,818 個唯一列全部已存在於第一張），但因為副本被動過、
不再逐格相同，先前的「sha256 完全相等」規則失效 → 約 4,800 筆重複寫入 raw_traces。
raw_traces 沒有內容層級唯一索引，DB 不會擋，結果是**事件頻次被灌水**，直接扭曲
軌跡密度與停留時間的判讀。

本規則的四個關鍵設計（各有測試對應）：
  1. fingerprint 建在 **_normalize_row 之後**，只取證據語意欄位 → 樣式／顏色／
     註解／未映射的標記欄天然不參與比較（副本被塗色仍可辨識為子集）。
  2. 用 **multiset（Counter）包含**而非 set：對每個 fingerprint 都要求
     count_B <= count_A，否則「A 有 1 筆、B 有 3 筆」會被誤判成完全重複而漏存
     2 筆真實事件。
  3. 只有 **new_rows == 0** 才整張跳過；有任何新列就整張保留（寧可多存不可漏存）。
  4. 去重狀態是 _iter_rows_excel 的**區域變數**，每次呼叫重建 → **不可能跨檔案
     去重**（不同單位交付的同批資料是各自獨立的證物）。

真實樣本檔在 .gitignore 內（含真實個資），CI 取不到 → 相關測試以 skipif 條件跳過。
"""
from __future__ import annotations

import io
import os
from typing import Any, Dict, List

import pytest

os.environ.setdefault("DATABASE_URL", "postgresql://user:pass@localhost:5432/fakedb")
os.environ.setdefault("SECRET_KEY", "test-secret-key-only-for-pytest")
os.environ.setdefault("AUTH_ENABLED", "true")

SAMPLE_DIR = "/Users/chenguanjun/Desktop/Python程序開發/CellTrail/基地台位置範例檔案"
_has_samples = os.path.isdir(SAMPLE_DIR)
needs_samples = pytest.mark.skipif(
    not _has_samples, reason="真實樣本檔不在版控內（含個資），此環境取不到"
)


# ─────────────────────────────────────────────────────────────
# 共用工具
# ─────────────────────────────────────────────────────────────
def _make_xlsx(sheets: Dict[str, List[List[Any]]]) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for r in rows:
            ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _patch_active_map_to_default(monkeypatch):
    import app.services.carrier_profile as cp

    monkeypatch.setattr(cp, "_load_default_profile_from_db", lambda: None)
    cp.invalidate_cache()


def _capture_skips(monkeypatch) -> List[Dict[str, Any]]:
    """攔截結構化 log，取得跳過分頁的稽核紀錄。"""
    import app.services.ingest as ing

    got: List[Dict[str, Any]] = []

    def _fake(event, **fields):
        if event == "ingest.sheet.skipped":
            got.append(dict(fields))

    monkeypatch.setattr(ing, "log_info", _fake)
    return got


HEADER = ["進入基地台時間", "基地台停留時間", "離開基地台時間",
          "離開基地台編號", "離開基地台地址"]


def _row(day: int, cell: str = "466970829017141", addr: str = "高雄市湖內區中山路２段290號"):
    return [f"2026-06-{day:02d}T10:00:00", "2700", f"2026-06-{day:02d}T10:45:00", cell, addr]


def _sheet(days) -> List[List[Any]]:
    return [HEADER] + [_row(d) for d in days]


def _yield_count(blob: bytes) -> int:
    from app.services.ingest import _iter_rows_excel

    return len(list(_iter_rows_excel(blob)))


# ─────────────────────────────────────────────────────────────
# 1. 完全相同的分頁
# ─────────────────────────────────────────────────────────────
def test_identical_sheet_is_skipped(monkeypatch):
    _patch_active_map_to_default(monkeypatch)
    skips = _capture_skips(monkeypatch)

    rows = _sheet([1, 2, 3])
    blob = _make_xlsx({"資料": rows, "副本": [r[:] for r in rows]})
    assert _yield_count(blob) == 3

    assert len(skips) == 1
    s = skips[0]
    assert s["reason"] == "subset_duplicate_sheet"
    assert s["reference_sheet"] == "sheet#0"
    assert s["sheet"] == "sheet#1"
    assert s["valid_rows"] == 3 and s["duplicate_rows"] == 3 and s["new_rows"] == 0


# ─────────────────────────────────────────────────────────────
# 2. 純子集分頁（031543 的真實情境）
# ─────────────────────────────────────────────────────────────
def test_pure_subset_sheet_is_skipped(monkeypatch):
    """B 是 A 的純子集（少了幾列）→ 整張跳過。"""
    _patch_active_map_to_default(monkeypatch)
    skips = _capture_skips(monkeypatch)

    blob = _make_xlsx({"全量": _sheet([1, 2, 3, 4, 5]), "標了重點的副本": _sheet([2, 4])})
    assert _yield_count(blob) == 5, "只應保留全量分頁的 5 列"
    assert skips[0]["reason"] == "subset_duplicate_sheet"
    assert skips[0]["valid_rows"] == 2 and skips[0]["new_rows"] == 0


def test_subset_rule_ignores_non_evidence_columns(monkeypatch):
    """
    副本被加上標記欄／改了未映射欄，仍應判定為子集。
    這是「fingerprint 建在 normalize 之後」的直接驗證：未被 header_map 認識的欄
    不會進入 normalized row，也就不可能影響比對。
    """
    _patch_active_map_to_default(monkeypatch)
    _capture_skips(monkeypatch)

    a = _sheet([1, 2, 3])
    b = [HEADER + ["承辦註記", "已複核"]] + [_row(d) + ["重點", "V"] for d in (1, 2, 3)]
    assert _yield_count(_make_xlsx({"A": a, "B": b})) == 3


# ─────────────────────────────────────────────────────────────
# 3. multiset：B 的某 fingerprint 數量超過 A → 不可跳過
# ─────────────────────────────────────────────────────────────
def test_multiset_count_exceeding_is_not_skipped(monkeypatch):
    """
    A 有 1 筆 X，B 有 3 筆 X。用 set 比對會誤判「B ⊆ A」而整張丟掉，
    漏掉 2 筆真實的獨立事件；multiset 必須把超出的量算成 new_rows。
    """
    _patch_active_map_to_default(monkeypatch)
    skips = _capture_skips(monkeypatch)

    a = [HEADER, _row(1)]
    b = [HEADER, _row(1), _row(1), _row(1)]
    assert _yield_count(_make_xlsx({"A": a, "B": b})) == 1 + 3, "超量的重複列必須保留"
    assert not [s for s in skips if s["reason"] == "subset_duplicate_sheet"]


def test_equal_counts_are_subset(monkeypatch):
    """count_B == count_A 屬於「涵蓋」，仍應跳過（邊界）。"""
    _patch_active_map_to_default(monkeypatch)
    _capture_skips(monkeypatch)

    a = [HEADER, _row(1), _row(1)]
    b = [HEADER, _row(1), _row(1)]
    assert _yield_count(_make_xlsx({"A": a, "B": b})) == 2


# ─────────────────────────────────────────────────────────────
# 4. 99% 重複 + 1 筆新列 → 整張保留
# ─────────────────────────────────────────────────────────────
def test_almost_duplicate_with_one_new_row_is_kept(monkeypatch):
    _patch_active_map_to_default(monkeypatch)
    skips = _capture_skips(monkeypatch)

    a = _sheet(range(1, 101))                      # 100 列
    b = _sheet(range(1, 101)) + [_row(28, cell="466970829099999")]   # 100 重複 + 1 新
    assert _yield_count(_make_xlsx({"A": a, "B": b})) == 100 + 101, \
        "只要有 1 列是新的，整張分頁都必須保留"
    assert not [s for s in skips if s["reason"] == "subset_duplicate_sheet"]


# ─────────────────────────────────────────────────────────────
# 5. 不得跨檔案去重
# ─────────────────────────────────────────────────────────────
def test_no_cross_file_deduplication(monkeypatch):
    """同一份內容放在兩個不同檔案 → 各自完整解析，互不影響。"""
    _patch_active_map_to_default(monkeypatch)
    _capture_skips(monkeypatch)

    blob = _make_xlsx({"資料": _sheet([1, 2, 3])})
    assert _yield_count(blob) == 3
    assert _yield_count(blob) == 3, "第二次呼叫不得沿用上一次的去重狀態"
    assert _yield_count(_make_xlsx({"另一個檔": _sheet([1, 2, 3])})) == 3


# ─────────────────────────────────────────────────────────────
# 6. 稽核與隱私要求
# ─────────────────────────────────────────────────────────────
def test_skip_record_is_auditable_and_pii_free(monkeypatch):
    """
    跳過必須留下可稽核欄位，且不得含分頁名稱／儲存格內容。
    分頁名實測可能直接是門號或對象姓名（如「0958549697 雙向歷程（嫌1）」）。
    """
    _patch_active_map_to_default(monkeypatch)
    skips = _capture_skips(monkeypatch)

    rows = _sheet([1, 2])
    _yield_count(_make_xlsx({
        "0958549697 雙向歷程（嫌1）": rows,
        "0958549697 副本 王小明": [r[:] for r in rows],
    }))

    assert len(skips) == 1
    s = skips[0]
    for k in ("reason", "reference_sheet", "valid_rows", "duplicate_rows", "new_rows"):
        assert k in s, f"稽核欄位缺少 {k}"
    blob = repr(s)
    assert "0958549697" not in blob and "王小明" not in blob, "log 不得含 PII"
    assert "高雄市" not in blob and "466970829017141" not in blob, "log 不得含證物內容"
    assert s["sheet"] == "sheet#1" and s["reference_sheet"] == "sheet#0"


def test_fingerprint_excludes_invalid_rows(monkeypatch):
    """
    缺時間／缺定位資訊的雜訊列不計入比對（下游本來就會 skip）。
    否則兩張分頁之間的雜訊列數差異會干擾子集判定。
    """
    from app.services.ingest import _evidence_fingerprint

    assert _evidence_fingerprint({"離開基地台編號": "X1"}) is None, "缺時間 → 非有效列"
    assert _evidence_fingerprint({"進入基地台時間": "2026-06-01T10:00:00"}) is None, \
        "缺 cell_id/地址/座標 → 非有效列"
    assert _evidence_fingerprint({
        "進入基地台時間": "2026-06-01T10:00:00", "離開基地台編號": "X1",
    }) is not None


def test_fingerprint_is_type_canonical():
    """
    同一筆證據在不同分頁可能一邊存成 datetime、一邊存成字串；
    不先 canonical 化會產生不同 fingerprint，子集判定就會失效。
    """
    from datetime import datetime

    from app.services.ingest import _evidence_fingerprint

    a = _evidence_fingerprint({
        "進入基地台時間": "2026-06-01T10:00:00", "離開基地台編號": "X1"})
    b = _evidence_fingerprint({
        "進入基地台時間": datetime(2026, 6, 1, 10, 0, 0), "離開基地台編號": "X1"})
    assert a == b


# ─────────────────────────────────────────────────────────────
# 7. 真實樣本回歸（檔案不在版控內 → 條件跳過）
# ─────────────────────────────────────────────────────────────
def _sample(name: str) -> bytes:
    return open(os.path.join(SAMPLE_DIR, name), "rb").read()


@needs_samples
def test_regression_031543_subset_sheet(monkeypatch):
    """031543：第二分頁為純子集，須跳過且不得重複寫入。"""
    fn = "031543 蘇網路 1150601-0711(0708基地台).xlsx"
    if not os.path.exists(os.path.join(SAMPLE_DIR, fn)):
        pytest.skip("樣本檔不存在")
    _patch_active_map_to_default(monkeypatch)
    skips = _capture_skips(monkeypatch)

    n = _yield_count(_sample(fn))
    subset = [s for s in skips if s["reason"] == "subset_duplicate_sheet"]
    assert len(subset) == 1 and subset[0]["new_rows"] == 0
    assert n < 6000, f"子集分頁未被跳過（yield={n}，翻倍即回歸）"


@needs_samples
def test_regression_029935_identical_sheet(monkeypatch):
    """複本 029935：兩分頁逐格相同，須維持既有正確行為。"""
    fn = "複本 029935 陳1號機網路 1150601-0711(0708基地台).xlsx"
    if not os.path.exists(os.path.join(SAMPLE_DIR, fn)):
        pytest.skip("樣本檔不存在")
    _patch_active_map_to_default(monkeypatch)
    skips = _capture_skips(monkeypatch)

    n = _yield_count(_sample(fn))
    subset = [s for s in skips if s["reason"] == "subset_duplicate_sheet"]
    assert len(subset) == 1 and subset[0]["new_rows"] == 0
    assert n < 6000, f"重複分頁未被跳過（yield={n}）"


@needs_samples
def test_regression_other_samples_have_no_unexpected_dedup(monkeypatch):
    """
    其餘樣本檔不得出現任何子集跳過 —— 這是「規則只命中副本分頁、
    不誤殺真實多分頁資料」的守門（周蔓達 13 個月分頁、電話通聯 6 個對象分頁）。
    刻意斷言「行為」而非「筆數」：樣本檔不在版控內、內容可能被承辦人更新，
    綁死筆數只會製造假警報。
    """
    _patch_active_map_to_default(monkeypatch)
    known_dup = {"031543", "複本 029935"}
    checked = 0
    for fn in sorted(os.listdir(SAMPLE_DIR)):
        if not fn.lower().endswith((".xlsx", ".xltx")) or fn.startswith("."):
            continue
        # `~$檔名.xlsx` 是 Excel 開檔時產生的鎖定暫存檔（約 165 bytes，非資料）。
        # 承辦人只要正開著任一樣本檔就會出現，pandas 讀它會拋
        # 「Excel file format cannot be determined」→ 測試偽失敗。
        if fn.startswith("~$"):
            continue
        if any(fn.startswith(k) for k in known_dup):
            continue
        skips = _capture_skips(monkeypatch)
        n = _yield_count(_sample(fn))
        assert n > 0, f"{fn[:12]}… 解析不出任何列"
        subset = [s for s in skips if s["reason"] == "subset_duplicate_sheet"]
        assert not subset, f"{fn[:12]}… 出現非預期的子集跳過：{subset}"
        checked += 1
    assert checked >= 10, f"實際只檢查了 {checked} 個樣本檔"
